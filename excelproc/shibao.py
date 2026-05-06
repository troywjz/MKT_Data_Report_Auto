# 时报数据模块
# 抓取百度、360、必应当天数据，写入02时报新增.xlsx

import sys
import os
import glob
import pandas as pd
from datetime import datetime, timedelta, time
from openpyxl import load_workbook, Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from spider.baidudata import baidudata
from spider.slldata import slldata
from spider.bingdata import bingdata

# 配置常量
SHIBAO_FILE = "02时报新增.xlsx"
SHIBAO_SHEET = "时报底表"
TEMPLATE_FILE = "数据模版.xlsx"
DATAS_DIR = "datas"

# 平台中文名映射
PLATFORM_NAME_MAP = {
    "baidu": "百度",
    "360": "360",
    "bing": "必应"
}


def get_closest_half_hour():
    """获取离当前时间最接近的整点半点时间"""
    now = datetime.now()
    minute = now.minute
    hour = now.hour

    if minute < 15:
        target_minute = 0
    elif minute < 45:
        target_minute = 30
    else:
        if hour == 23:
            target_minute = 30
        else:
            target_minute = 0
            hour = hour + 1

    return time(hour, target_minute, 0)


def load_template_mapping():
    """
    加载日报模板的"时报"sheet，建立账户名 -> (D列项目, E列渠道, G列返点) 的映射
    返回: dict {账户名: {'D': 项目, 'E': 渠道, 'G': 返点}}
    """
    try:
        wb = load_workbook(TEMPLATE_FILE, data_only=True)
        ws = wb['时报']

        mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            account = row[5]  # F列 = index 5 (账户名)
            if account:
                mapping[account] = {
                    'D': row[3],  # 项目
                    'E': row[4],  # 渠道
                    'G': row[6],  # 返点
                }

        wb.close()
        return mapping
    except Exception as e:
        print(f"[!] 加载时报模板失败: {e}")
        return {}


def load_lead_rules():
    """
    加载"时报线索识别方式"sheet，建立账户名 -> 读码方式 的映射
    返回: dict {账户名: 读码方式}
    """
    try:
        wb = load_workbook(TEMPLATE_FILE, data_only=True)
        ws = wb['时报线索识别方式']

        mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[2]:  # B列=账户名, C列=读码方式
                account = row[1]
                pattern = row[2]
                mapping[account] = pattern

        wb.close()
        return mapping
    except Exception as e:
        print(f"[!] 加载线索匹配规则失败: {e}")
        return {}


def count_leads_by_account(crm_df, lead_rules):
    """
    统计每个账户的线索数量
    crm_df: CRM DataFrame (包含 site_id 列，即L列)
    lead_rules: {账户名: 读码方式}
    返回: dict {账户名: 线索数}
    """
    try:
        if crm_df is None or crm_df.empty:
            print("[!] CRM 数据为空")
            return {account: 0 for account in lead_rules}

        lead_count = {account: 0 for account in lead_rules}

        # 遍历每一行，检查 site_id 是否包含账户的读码方式
        for _, row in crm_df.iterrows():
            site_id = row.get('site_id', '')
            if pd.isna(site_id) or site_id == '':
                continue
            site_id_str = str(site_id)
            for account, pattern in lead_rules.items():
                if pattern and pattern in site_id_str:
                    lead_count[account] = lead_count.get(account, 0) + 1

        return lead_count
    except Exception as e:
        print(f"[!] 统计线索失败: {e}")
        return {account: 0 for account in lead_rules}


def write_to_shibao(data_list, lead_counts, template_mapping):
    """
    将数据写入时报Excel (02时报新增.xlsx)
    data_list: 包含字典的列表
    lead_counts: {账户名: 线索数}
    template_mapping: 模板映射 {账户名: {'D': 项目, 'E': 渠道, 'G': 返点}}
    """
    date_obj = datetime.now()
    year = date_obj.year
    date_only = datetime(date_obj.year, date_obj.month, date_obj.day)
    time_half_hour = get_closest_half_hour()

    try:
        wb = load_workbook(SHIBAO_FILE)
        ws = wb[SHIBAO_SHEET]
    except Exception as e:
        print(f"[!] 读取时报文件失败: {e}")
        return 0

    count = 0
    skipped = 0
    for row_data in data_list:
        platform = row_data.get("platform", "")
        account = row_data.get("account", "")
        cost = float(row_data.get("cost", 0))
        show = int(row_data.get("show", 0))
        click = int(row_data.get("click", 0))

        # 检查账户是否在模板中，不在则跳过
        if account not in template_mapping:
            skipped += 1
            continue

        # 从模板获取 D, E, G 列
        template_info = template_mapping.get(account, {})
        d_value = template_info.get('D', '')  # 项目
        e_value = template_info.get('E', '')  # 渠道
        g_value = template_info.get('G', 0)   # 返点

        # 计算 H列 = K / (1 + G)
        if g_value and isinstance(g_value, (int, float)) and g_value > 0:
            h_value = cost / (1 + g_value)
        else:
            h_value = 0

        # 获取线索数
        l_value = lead_counts.get(account, 0)

        # 构建行数据
        # A,B,C, D项目, E渠道, F账户名, G返点, H实际消费, I展现, J点击, K消费, L名片数
        row = [
            year,           # A列: 年份
            date_only,      # B列: 日期
            time_half_hour, # C列: 时间
            d_value,        # D列: 项目
            e_value,        # E列: 渠道
            account,        # F列: 账户名
            g_value,        # G列: 返点
            round(h_value, 2),  # H列: 实际消费
            show,           # I列: 展现
            click,           # J列: 点击
            cost,           # K列: 消费
            l_value,        # L列: 名片数(线索)
        ]

        ws.append(row)
        count += 1

    try:
        wb.save(SHIBAO_FILE)
        wb.close()
        print(f"[√] 时报数据已保存，共 {count} 条")
        return count
    except Exception as e:
        print(f"[!] 保存时报文件失败: {e}")
        return 0


def run_crawler(platform, date_input):
    """运行指定平台的爬虫"""
    try:
        if platform == "baidu":
            data = baidudata(date_input)
        elif platform == "360":
            data = slldata(date_input)
        elif platform == "bing":
            data = bingdata(date_input)
        else:
            print(f"[!] 不支持的平台: {platform}")
            return []

        for item in data:
            item["platform"] = platform

        return data

    except Exception as e:
        print(f"[!] {platform} 数据抓取失败: {e}")
        return []


def check_existing_shibao(date_obj, time_half_hour):
    """
    检查时报文件中是否已有指定日期和时间的记录
    返回: True/False
    """
    if not os.path.exists(SHIBAO_FILE):
        return False

    try:
        wb = load_workbook(SHIBAO_FILE, data_only=True)
        ws = wb[SHIBAO_SHEET]

        for row in ws.iter_rows(min_row=2, values_only=True):
            b_val = row[1]  # B列: 日期
            c_val = row[2]  # C列: 时间

            if b_val is None or c_val is None:
                continue

            row_date = None
            if isinstance(b_val, datetime):
                row_date = b_val.date()
            else:
                continue

            row_time = None
            if isinstance(c_val, datetime):
                row_time = c_val.time()
            elif isinstance(c_val, time):
                row_time = c_val
            else:
                continue

            if row_date == date_obj.date() and row_time == time_half_hour:
                wb.close()
                return True

        wb.close()
        return False
    except Exception as e:
        print(f"[!] 检查时报文件失败: {e}")
        return False


def read_existing_shibao(date_obj, time_half_hour):
    """
    从时报文件读取指定日期和时间的已有数据
    返回: data_list
    """
    if not os.path.exists(SHIBAO_FILE):
        return []

    try:
        wb = load_workbook(SHIBAO_FILE, data_only=True)
        ws = wb[SHIBAO_SHEET]

        data_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            b_val = row[1]
            c_val = row[2]
            f_val = row[5]  # F列: 账户名
            i_val = row[8]  # I列: 展现
            j_val = row[9]  # J列: 点击
            k_val = row[10]  # K列: 消费

            if b_val is None or c_val is None:
                continue

            row_date = None
            if isinstance(b_val, datetime):
                row_date = b_val.date()
            else:
                continue

            row_time = None
            if isinstance(c_val, datetime):
                row_time = c_val.time()
            elif isinstance(c_val, time):
                row_time = c_val
            else:
                continue

            if row_date == date_obj.date() and row_time == time_half_hour:
                data_list.append({
                    "date": date_obj.strftime("%Y-%m-%d"),
                    "account": f_val,
                    "device": "",
                    "show": int(i_val) if i_val else 0,
                    "click": int(j_val) if j_val else 0,
                    "cost": float(k_val) if k_val else 0.0,
                    "platform": "",
                })

        wb.close()
        return data_list
    except Exception as e:
        print(f"[!] 读取时报文件失败: {e}")
        return []


def update_l_column(file_path, sheet_name, start_row, lead_counts):
    """
    更新指定行范围的L列（线索数列）
    """
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        for row_idx in range(start_row, ws.max_row + 1):
            account = ws.cell(row=row_idx, column=6).value
            if account and account in lead_counts:
                ws.cell(row=row_idx, column=12).value = lead_counts[account]

        wb.save(file_path)
        wb.close()
    except Exception as e:
        print(f"[!] 更新L列失败: {e}")


def update_shbaotemplate_from_02(today_date, time_half_hour):
    """
    从02时报新增.xlsx读取今日、昨日、上周同期的数据
    按账户名匹配，写入数据模版.xlsx的时报发送模板
    """
    TARGET_SHEET = "时报发送模板"

    # 读取02时报新增数据
    try:
        wb_src = load_workbook(SHIBAO_FILE, data_only=True)
        ws_src = wb_src[SHIBAO_SHEET]
    except Exception as e:
        print(f"[!] 读取02时报新增失败: {e}")
        return

    all_rows = list(ws_src.iter_rows(min_row=2, values_only=True))
    wb_src.close()

    if not all_rows:
        print("[!] 02时报新增无数据")
        return

    today = today_date.date()
    time_slots_today = set()
    for row in all_rows:
        if row[1] and row[1].date() == today:
            time_slots_today.add(str(row[2]))
    if not time_slots_today:
        print("[!] 今日无数据，无法确定时间点")
        return
    target_time = sorted(time_slots_today, reverse=True)[0]

    yesterday_date = today - timedelta(days=1)
    week_ago_date = today - timedelta(days=7)

    def sum_by_account(date, time_slot):
        result = {}
        for row in all_rows:
            row_date = row[1].date() if row[1] else None
            row_time = row[2]
            if row_date == date and str(row_time) == str(time_slot):
                account = row[5]
                if account:
                    if account not in result:
                        result[account] = {'展现': 0, '点击': 0, '消费': 0, '实际消费': 0, '线索': 0}
                    result[account]['展现'] += int(row[8] or 0)
                    result[account]['点击'] += int(row[9] or 0)
                    result[account]['消费'] += float(row[10] or 0)
                    result[account]['实际消费'] += float(row[7] or 0)
                    result[account]['线索'] += int(row[11] or 0)
        return result

    def normalize_channel(ch):
        if ch == '三六零' or ch == '360' or ch == 360:
            return '三六零'
        if ch == '百度' or ch == '百度':
            return '百度'
        if ch == '必应' or ch == '必应':
            return '必应'
        if ch == '神马':
            return '神马'
        if ch == '广点通':
            return '广点通'
        return str(ch)

    account_channel_map = {}
    for row in all_rows:
        if row[1] and row[1].date() == today and str(row[2]) == str(target_time):
            account = row[5]
            channel = row[4]
            if account and channel:
                account_channel_map[account] = normalize_channel(channel)

    try:
        wb_dst = load_workbook(TEMPLATE_FILE)
        ws_dst = wb_dst[TARGET_SHEET]
        # B1 写入时报日期+时间
        ws_dst.cell(row=1, column=1).value = today_date.date() if hasattr(today_date, 'date') else today_date
        ws_dst.cell(row=1, column=2).value = time_half_hour
    except Exception as e:
        print(f"[!] 打开时报发送模板失败: {e}")
        return

    def write_period_rows(start_row, end_row, data_dict, label):
        updated = 0
        for row_idx in range(start_row, end_row + 1):
            period_cell = ws_dst.cell(row=row_idx, column=1).value
            account_cell = ws_dst.cell(row=row_idx, column=3).value
            channel = ws_dst.cell(row=row_idx, column=2).value

            if str(period_cell) != label:
                continue

            is_total = (str(account_cell) == '合计' or str(channel) == '合计')

            def safe_div(a, b, mul=1):
                try:
                    if b == 0:
                        return 0
                    return round(a / b * mul, 2)
                except:
                    return 0

            if account_cell and account_cell in data_dict:
                d = data_dict[account_cell]
                lead = int(d['线索'])
                show = int(d['展现'])
                click = int(d['点击'])
                cost = round(float(d['消费']), 2)
                actual = round(float(d['实际消费']), 2)
            elif is_total:
                if channel and str(channel) != '合计':
                    channel_norm = normalize_channel(channel)
                    totals = {'展现': 0, '点击': 0, '消费': 0, '实际消费': 0, '线索': 0}
                    for acc, d in data_dict.items():
                        acc_ch_norm = normalize_channel(account_channel_map.get(acc, ''))
                        if acc_ch_norm == channel_norm:
                            totals['展现'] += d['展现']
                            totals['点击'] += d['点击']
                            totals['消费'] += d['消费']
                            totals['实际消费'] += d['实际消费']
                            totals['线索'] += d['线索']
                    lead = int(totals['线索'])
                    show = int(totals['展现'])
                    click = int(totals['点击'])
                    cost = round(totals['消费'], 2)
                    actual = round(totals['实际消费'], 2)
                else:
                    totals = {'展现': 0, '点击': 0, '消费': 0, '实际消费': 0, '线索': 0}
                    for d in data_dict.values():
                        totals['展现'] += d['展现']
                        totals['点击'] += d['点击']
                        totals['消费'] += d['消费']
                        totals['实际消费'] += d['实际消费']
                        totals['线索'] += d['线索']
                    lead = int(totals['线索'])
                    show = int(totals['展现'])
                    click = int(totals['点击'])
                    cost = round(totals['消费'], 2)
                    actual = round(totals['实际消费'], 2)
            else:
                lead = show = click = 0
                cost = actual = 0.0

            actual_cost = safe_div(actual, lead)
            card_cost = safe_div(cost, lead)
            cpc = safe_div(cost, click)
            ctr = safe_div(click, show)
            cvr = safe_div(lead, click)
            cpm = safe_div(cost, show, 1000)

            ws_dst.cell(row=row_idx, column=4).value = lead
            ws_dst.cell(row=row_idx, column=5).value = actual_cost
            ws_dst.cell(row=row_idx, column=6).value = card_cost
            ws_dst.cell(row=row_idx, column=7).value = show
            ws_dst.cell(row=row_idx, column=8).value = click
            ws_dst.cell(row=row_idx, column=9).value = cost
            ws_dst.cell(row=row_idx, column=10).value = actual
            ws_dst.cell(row=row_idx, column=11).value = cpc
            ws_dst.cell(row=row_idx, column=12).value = ctr
            ws_dst.cell(row=row_idx, column=13).value = cvr
            ws_dst.cell(row=row_idx, column=14).value = cpm

            updated += 1

        return updated

    data_today = sum_by_account(today, target_time)
    data_yesterday = sum_by_account(yesterday_date, target_time)
    data_week_ago = sum_by_account(week_ago_date, target_time)

    write_period_rows(3, 12, data_week_ago, '上周同期')
    write_period_rows(13, 22, data_yesterday, '昨日同期')
    write_period_rows(23, 32, data_today, '今日')

    try:
        wb_dst.save(TEMPLATE_FILE)
        wb_dst.close()
        print(f"[√] 时报发送模板已更新")
    except Exception as e:
        print(f"[!] 保存时报发送模板失败: {e}")


def run():
    """时报主函数"""
    print("时报数据抓取开始")

    date_obj = datetime.now()
    time_half_hour = get_closest_half_hour()
    today_date = datetime(date_obj.year, date_obj.month, date_obj.day)

    print(f"[-] 目标日期: {today_date.date()}")
    print(f"[-] 目标时间: {time_half_hour}")

    existing = check_existing_shibao(today_date, time_half_hour)
    platforms = ["baidu", "360", "bing"]

    if existing:
        print(f"[*] 检测到 {today_date.date()} {time_half_hour} 已有数据，跳过抓取")
        all_data = read_existing_shibao(today_date, time_half_hour)
        print(f"[*] 读取到 {len(all_data)} 条现有数据")
        start_row_before = 0
        skip_crawl = True
    else:
        print(f"[*] 未检测到现有数据，开始抓取...")
        all_data = []
        for platform in platforms:
            data = run_crawler(platform, "shibao")
            all_data.extend(data)

        if not all_data:
            print("[!] 没有抓取到数据")
            return

        template_mapping = load_template_mapping()

        try:
            wb_temp = load_workbook(SHIBAO_FILE)
            ws_temp = wb_temp[SHIBAO_SHEET]
            start_row_before = ws_temp.max_row
            wb_temp.close()
        except:
            start_row_before = 0

        count = write_to_shibao(all_data, {}, template_mapping)

        if count == 0:
            print("[!] 没有数据写入")
            return

        start_row = start_row_before + 1
        skip_crawl = False

    template_mapping = load_template_mapping()

    lead_counts = {}
    if not skip_crawl:
        try:
            from spider.crmdata import crmdata as run_crm
            crm_df = run_crm("shibao", return_df=True)
        except Exception as e:
            print(f"[!] CRM数据获取失败: {e}")
            crm_df = None

        lead_rules = load_lead_rules()
        filtered_lead_rules = {acc: lead_rules[acc] for acc in template_mapping.keys() if acc in lead_rules}

        if crm_df is not None and not crm_df.empty and filtered_lead_rules:
            lead_counts = count_leads_by_account(crm_df, filtered_lead_rules)
        else:
            print("[!] 无法获取线索数据")

        if lead_counts and start_row_before > 0:
            update_l_column(SHIBAO_FILE, SHIBAO_SHEET, start_row_before + 1, lead_counts)

    update_shbaotemplate_from_02(today_date, time_half_hour)

    print("[√] 时报数据抓取完成")

    print_as_markdown()


def print_as_markdown():
    """读取时报发送模板，打印为markdown表格"""
    import gc
    gc.collect()
    wb = None
    try:
        wb = load_workbook(TEMPLATE_FILE, data_only=True)
        ws = wb['时报发送模板']

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return

        # 收集每列最大宽度
        col_widths = [0] * len(rows[0])
        for row in rows:
            for i, cell in enumerate(row):
                val = str(cell) if cell is not None else ""
                col_widths[i] = max(col_widths[i], len(val))

        # 打印表头
        header = rows[0]
        print("\n# 请openclaw发送以下数据作为时报 #\n")
        print("| " + " | ".join(str(h) if h is not None else "" for h in header) + " |")
        print("| " + " | ".join("-" * w for w in col_widths) + " |")

        # 打印数据行
        for row in rows[1:]:
            print("| " + " | ".join(str(c) if c is not None else "" for c in row) + " |")

        print()
    except Exception as e:
        print(f"[!] 读取时报发送模板失败: {e}")
    finally:
        if wb is not None:
            try:
                wb.close()
            except:
                pass
        gc.collect()


if __name__ == "__main__":
    run()
