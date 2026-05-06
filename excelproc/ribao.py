# 日报数据模块
# 抓取百度、360、必应指定日期数据，写入01日报新增.xlsx

import argparse
import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from spider.baidudata import baidudata
from spider.slldata import slldata
from spider.bingdata import bingdata

RIBAO_NEW_FILE = "01日报新增.xlsx"
RIBAO_SHEET = "数据底表-项目表（删实际消费）"
TEMPLATE_FILE = "数据模版.xlsx"


def parse_arguments():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description="日报数据抓取")
    parser.add_argument("-d", "--date", type=str, required=True, help="指定日期，格式: YYYY-MM-DD")
    return parser.parse_args()


def load_ribao_template():
    """加载日报模板，建立 (账户, 端口) -> (E-R列数据) 的映射"""
    try:
        wb = load_workbook(TEMPLATE_FILE, data_only=True)
        ws = wb['日报']

        mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[14]:
                account = row[14]
                port = row[16]
                if account and port:
                    template_row = {
                        'E': row[4],
                        'F': row[5],
                        'G': row[6],
                        'H': row[7],
                        'I': row[8],
                        'J': row[9],
                        'K': row[10],
                        'L': row[11],
                        'M': row[12],
                        'N': row[13],
                        'O': row[14],
                        'P': row[15],
                        'Q': row[16],
                        'R': row[17],
                    }
                    mapping[(account, port)] = template_row

        wb.close()
        print(f"[-] 已加载日报模板，共 {len(mapping)} 条映射")
        return mapping
    except Exception as e:
        print(f"[!] 加载日报模板失败: {e}")
        return {}


def calc_season(date_obj):
    """根据日期计算季节"""
    day = date_obj.day
    month = date_obj.month
    if day <= 8:
        return f"{month}月春季"
    elif day <= 15:
        return f"{month}月夏季"
    elif day <= 23:
        return f"{month}月秋季"
    else:
        return f"{month}月冬季"


def load_lead_rules():
    """
    加载"日报线索识别方式"sheet，建立 (账户, 端口) -> 识别方式 的映射
    """
    try:
        wb = load_workbook(TEMPLATE_FILE, data_only=True)
        ws = wb['日报线索识别方式']

        mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            account = row[2]
            port = row[3]
            pattern = row[5]
            if account and port and pattern:
                mapping[(account, port)] = pattern

        wb.close()
        print(f"[-] 已加载线索识别规则，共 {len(mapping)} 条")
        return mapping
    except Exception as e:
        print(f"[!] 加载线索识别规则失败: {e}")
        return {}


def count_leads_by_account(crm_df, lead_rules):
    """
    统计每个(账户, 端口)的线索数量
    """
    try:
        if crm_df is None or crm_df.empty:
            print("[!] CRM 数据为空")
            return {key: 0 for key in lead_rules}

        lead_count = {key: 0 for key in lead_rules}

        site_id_col = None
        for col in crm_df.columns:
            if str(col).lower() == 'site_id':
                site_id_col = col
                break

        if site_id_col is None:
            print("[!] CRM 数据中未找到 site_id 列")
            return {key: 0 for key in lead_rules}

        for _, row in crm_df.iterrows():
            site_id = row.get(site_id_col, '')
            if pd.isna(site_id) or site_id == '':
                continue
            site_id_str = str(site_id)

            for (account, port), pattern in lead_rules.items():
                if pattern and pattern in site_id_str:
                    lead_count[(account, port)] = lead_count.get((account, port), 0) + 1

        return lead_count
    except Exception as e:
        print(f"[!] 统计线索失败: {e}")
        import traceback
        traceback.print_exc()
        return {key: 0 for key in lead_rules}


def write_to_ribao(data_list, date_str, lead_counts=None):
    """将数据写入日报Excel文件 (01日报新增.xlsx)"""
    if lead_counts is None:
        lead_counts = {}

    print(f"[-] 待写入数据条数: {len(data_list)}")
    if not data_list:
        print("[!] 没有数据需要写入")
        return 0

    template_mapping = load_ribao_template()

    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    year_str = f"{date_obj.year}年"
    month_first_day = datetime(date_obj.year, date_obj.month, 1)
    day_first = datetime(date_obj.year, date_obj.month, date_obj.day)
    season_str = calc_season(date_obj)

    rows_to_write = []
    not_found_accounts = set()

    for row_data in data_list:
        account = row_data.get("account", "")
        device = row_data.get("device", "")
        cost = float(row_data.get("cost", 0))
        show = int(row_data.get("show", 0))
        click = int(row_data.get("click", 0))

        template_row = template_mapping.get((account, device))
        if not template_row:
            not_found_accounts.add((account, device))
            print(f"[!] 未找到模板映射: 账户={account}, 端口={device}")
            continue

        rebate = template_row['R']
        if rebate and rebate > 0:
            actual_cost = cost / rebate
        else:
            actual_cost = 0

        lead_num = lead_counts.get((account, device), 0)

        rows_to_write.append({
            'account': account,
            'device': device,
            'cost': cost,
            'show': show,
            'click': click,
            'template': template_row,
            'actual_cost': actual_cost,
            'lead_num': lead_num,
        })

    if not rows_to_write:
        print("[!] 没有有效数据需要写入")
        return 0

    if not_found_accounts:
        print(f"[!] 共有 {len(not_found_accounts)} 个账户未找到模板映射")

    try:
        file_exists = os.path.exists(RIBAO_NEW_FILE)

        if file_exists:
            wb = load_workbook(RIBAO_NEW_FILE)
            ws = wb.active
            start_row = ws.max_row + 1
            # 第一条新数据行加粗
            first_data_row = start_row
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = RIBAO_SHEET
            # 写入表头并加粗
            headers = ['年份', '月初', '季节', '日期', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', '账户', '端口', 'Q', '返点', '展现', '点击', '消费', 'W', 'X', 'Y', '线索数', 'Z', '实际消费']
            ws.append(headers)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=1, column=col_idx).font = Font(bold=True)
            start_row = 2
            first_data_row = None  # 新文件不需要加粗数据行

        for idx, item in enumerate(rows_to_write):
            template = item['template']

            row = [
                year_str,
                month_first_day,
                season_str,
                day_first,
                template['E'],
                template['F'],
                template['G'],
                template['H'],
                template['I'],
                template['J'],
                template['K'],
                template['L'],
                template['M'],
                template['N'],
                template['O'],
                template['P'],
                template['Q'],
                template['R'],
                item['show'],
                item['click'],
                item['cost'],
                None,
                None,
                None,
                item['lead_num'],
                None,
                item['actual_cost'],
            ]
            ws.append(row)

            # 如果是追加到已有文件，第一条数据行加粗
            if first_data_row and idx == 0:
                for col_idx in range(1, len(row) + 1):
                    ws.cell(row=first_data_row, column=col_idx).font = Font(bold=True)

        wb.save(RIBAO_NEW_FILE)
        wb.close()
        print(f"[√] 日报数据已保存到: {RIBAO_NEW_FILE}，共 {len(rows_to_write)} 条")
        return len(rows_to_write)
    except Exception as e:
        print(f"[!] 保存日报文件失败: {e}")
        import traceback
        traceback.print_exc()
        return 0


def run_crawler(platform, date_input):
    """运行指定平台的爬虫"""
    print(f"\n{'=' * 50}")
    print(f"开始抓取 {platform} 数据...")
    print(f"{'=' * 50}")

    try:
        if platform == "baidu":
            data = baidudata(date_input)
        elif platform == "360":
            data = slldata(date_input)
        elif platform == "bing":
            data = bingdata(date_input)
        else:
            print(f"[!] 不支持的平台: {platform}")
            return []

        for item in data:
            item["platform"] = platform

        print(f"[√] {platform} 数据抓取完成，获取 {len(data)} 条数据")
        return data

    except Exception as e:
        print(f"[!] {platform} 数据抓取失败: {e}")
        return []


def run(date_str=None):
    """日报主函数"""
    print("=" * 50)
    print("日报数据抓取")
    print("=" * 50)

    if not date_str:
        args = parse_arguments()
        date_str = args.date
    else:
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            print(f"[!] 日期格式错误: {date_str}，应为 YYYY-MM-DD")
            sys.exit(1)

    print(f"[-] 目标日期: {date_str}")

    platforms = ["baidu", "360", "bing"]

    all_data = []
    for platform in platforms:
        data = run_crawler(platform, date_str)
        all_data.extend(data)

    print(f"\n{'=' * 50}")
    print("获取 CRM 线索数据...")
    print(f"{'=' * 50}")

    lead_counts = {}
    try:
        from spider.crmdata import crmdata as run_crm
        crm_df = run_crm(date_str, return_df=True)

        lead_rules = load_lead_rules()

        template_mapping = load_ribao_template()
        template_accounts = set(template_mapping.keys())
        filtered_lead_rules = {k: v for k, v in lead_rules.items() if k in template_accounts}

        if crm_df is not None and not crm_df.empty and filtered_lead_rules:
            lead_counts = count_leads_by_account(crm_df, filtered_lead_rules)
            print(f"[-] 线索统计完成")
        else:
            print("[!] 无法获取线索数据或无匹配规则")
    except Exception as e:
        print(f"[!] CRM 数据获取失败: {e}")

    print(f"\n{'=' * 50}")
    print("开始写入数据...")
    print(f"{'=' * 50}")

    count = write_to_ribao(all_data, date_str, lead_counts)

    print(f"\n{'=' * 50}")
    print(f"日报数据抓取完成！共 {count} 条数据")
    print(f"{'=' * 50}")


if __name__ == "__main__":
    run()

