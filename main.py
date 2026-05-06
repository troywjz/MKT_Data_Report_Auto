# 主程序 - 营销数据报告自动化
# 调度时报和日报程序完成数据抓取和写入

import argparse
import os
import sys
from datetime import datetime, timedelta
from dateutil.parser import parse as parse_date
from excelproc import shibao, ribao


RIBAO_NEW_FILE = "01日报新增.xlsx"


def normalize_date(date_str):
    """标准化日期格式，支持 2026-4-8、2026-4-08、2026-04-08 等格式"""
    try:
        dt = parse_date(date_str)
        return dt.strftime("%Y-%m-%d")
    except:
        return date_str


def get_latest_report_date():
    """读取日报输出文件最后30行的D列日期，返回最晚的一天"""
    if not os.path.exists(RIBAO_NEW_FILE):
        return None

    try:
        from openpyxl import load_workbook
        wb = load_workbook(RIBAO_NEW_FILE, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[-30:] if len(rows) > 30 else rows

        latest_date = None
        for row in data_rows:
            d_val = row[3]
            if d_val is None:
                continue
            if isinstance(d_val, datetime):
                d_date = d_val.date()
            else:
                try:
                    d_date = parse_date(str(d_val)).date()
                except:
                    continue

            if latest_date is None or d_date > latest_date:
                latest_date = d_date

        wb.close()
        return latest_date
    except Exception as e:
        print(f"[!] 读取日报文件失败: {e}")
        return None


def get_dates_to_process(latest_date):
    """根据最晚日报日期，计算需要做的日期范围: (latest_date + 1) -> yesterday"""
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)

    if latest_date is None:
        start_date = yesterday.replace(day=1)
    else:
        start_date = latest_date + timedelta(days=1)

    if start_date > yesterday:
        return []

    dates = []
    current = start_date
    while current <= yesterday:
        dates.append(current.strftime("%Y-%m-%d"))
        current += timedelta(days=1)

    return dates


def main():
    parser = argparse.ArgumentParser(description="营销数据报告自动化")
    parser.add_argument("mode", choices=["shibao", "ribao"], help="shibao: 时报(当天) | ribao: 日报")
    parser.add_argument("date", nargs="?", type=str, help="日报模式下指定日期，如: 2026-4-8")
    args = parser.parse_args()

    if args.mode == "shibao":
        shibao.run()

    elif args.mode == "ribao":
        if args.date:
            date_str = normalize_date(args.date)
            print(f"[=] 指定日期模式: {date_str}")
            ribao.run(date_str)
        else:
            print("[=] 无指定日期，自动模式")
            latest_date = get_latest_report_date()

            if latest_date is None:
                print("[-] 未找到历史日报，将从昨天开始制作")
            else:
                print(f"[-] 最后日报日期: {latest_date}")

            dates_to_do = get_dates_to_process(latest_date)

            if not dates_to_do:
                print("[*] 已完成所有日报，无需新增")
                return

            print(f"[-] 待做日报日期: {dates_to_do}")

            for date_str in dates_to_do:
                print(f"\n{'#' * 60}")
                print(f"# 开始制作 {date_str} 的日报")
                print(f"{'#' * 60}")
                ribao.run(date_str)

            print(f"\n{'=' * 60}")
            print(f"[√] 日报全部完成，共 {len(dates_to_do)} 天")
            print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
