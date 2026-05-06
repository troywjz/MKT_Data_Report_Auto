# -*- coding: utf-8 -*-
# 百度营销 - 客户中心 数据爬虫
# 功能：给定日期，返回数据列表
# 已打包成方法 baidudata("shibao")

# 调试时，打开最后几行的 if __name__ == "__main__": ...

import json
import time
import os
import datetime
from DrissionPage import ChromiumPage, ChromiumOptions
from DrissionPage.common import Keys
from openpyxl import Workbook, load_workbook
from app_config import (
    get_chromium_local_port,
    get_chromium_user_data_path,
    get_env,
    get_platform_credentials,
)

# ================= 配置区域 =================
CONFIG_FILE = "config.xlsx"
RESULT_FILE = "推广数据.xlsx"
TARGET_PLATFORM = "百度"

# 只有当 URL 包含这个时，才算真正进入后台
SUCCESS_URL_KEY = get_env("BAIDU_SUCCESS_URL_KEY")
LOGIN_URL_KEY = get_env("BAIDU_LOGIN_URL_KEY")
LOGIN_URL = get_env("BAIDU_LOGIN_URL")
FULL_TARGET_URL = get_env("BAIDU_TARGET_URL")


# date_input = "2025-11-24"
# date_input = "shibao"

# ===========================================

def switch_to_baidu_tab(page):
    """
    寻找并切换到包含 平台名 的标签页
    如果没有，则新建一个
    """
    # 尝试查找 URL 包含 keyword 的标签页
    target_keyword = "baidu"

    try:
        # 尝试查找标签页
        # 如果找不到，DrissionPage 会直接报错，跳转到 except
        tab = page.get_tab(url=target_keyword)

        print(f"[-] 检测到已存在的标签页: {tab.title}，正在激活...")
        tab.set.activate()
        return tab

    except Exception:
        # 捕获找不到标签页的错误，执行新建操作
        print("[-] 未检测到标签页，正在新建...")
        new_tab = page.new_tab(FULL_TARGET_URL)
        return new_tab

def get_credentials_from_excel():
    env_username, env_password = get_platform_credentials("BAIDU")
    if env_username:
        print("[-] Loaded Baidu credentials from environment")
        return env_username, env_password

    if not os.path.exists(CONFIG_FILE): return None, None
    try:
        wb = load_workbook(CONFIG_FILE, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row[0] and str(row[0]).strip() == TARGET_PLATFORM:
                print(f"[-] 读取到账号: {row[1]}")
                return str(row[1]), str(row[2])
        return None, None
    except:
        return None, None


def save_to_excel(data_rows, device_name="", data_date=""):
    try:
        if os.path.exists(RESULT_FILE):
            wb = load_workbook(RESULT_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["数据日期", "账户名称", "设备", "展现", "点击", "消费"])

        count = 0
        for row in data_rows:
            if int(row.get('impression', 0)) > 0:
                line = [
                    data_date,
                    row.get('userName'),
                    device_name,
                    int(row.get('impression')),
                    int(row.get('click', 0)),
                    float(row.get('cost'))
                ]
                ws.append(line)
                count += 1
        wb.save(RESULT_FILE)
        print(f"[√] 已将 {count} 条数据写入 {RESULT_FILE}")
    except Exception as e:
        print(f"[!] 写入 Excel 失败: {e}")


def wait_baidu_report_ready(page, timeout=30):
    """
    等待百度报表页面所有核心元素加载完成（与关系）
    核心元素：分页控件(条/页) + 推广设备选择器
    timeout: 最长等待时间(秒)，默认30秒
    """
    core_selectors = [
        'text:条/页',       # 分页控件
        'text:推广设备',    # 设备选择器
    ]

    try:
        page.wait.doc_loaded(timeout=10)
    except:
        pass

    start_time = time.time()
    found = set()

    while time.time() - start_time < timeout:
        for selector in core_selectors:
            if selector not in found:
                try:
                    ele = page.ele(selector, timeout=0.5)
                    if ele:
                        found.add(selector)
                        print(f"  [√] 已加载: {selector}")
                except:
                    pass

        if len(found) == len(core_selectors):
            elapsed = time.time() - start_time
            print(f"[-] 报表页面已就绪 (等待了 {elapsed:.1f} 秒)")
            time.sleep(0.5)
            return True

        time.sleep(0.5)

    print(f"[!] 等待超时({timeout}秒)，已加载: {[s for s in core_selectors if s in found]}")
    time.sleep(2)
    return False


def wait_page_stable(page):
    """通用等待函数（仅用于登录页等非报表场景）"""
    try:
        page.wait.doc_loaded(timeout=5)
    except:
        time.sleep(1)
    time.sleep(2)


def strict_login_process(page, username, password):
    """
    【严厉版】登录流程
    如果不进后台，死都不出来
    """
    wait_page_stable(page)

    print("\n" + "=" * 50)
    print("[-] 开始身份验证流程...")

    # 1. 判断当前状态
    # 如果不包含 关键词，说明没登录
    is_logged_in = (SUCCESS_URL_KEY in page.url) and (LOGIN_URL_KEY not in page.url)

    if is_logged_in:
        print("[√] 检测到已在报表后台，跳过登录。")
        return

    # 2. 进入登录页
    print("[-] 未登录，跳转至登录页...")
    page.get(FULL_TARGET_URL)
    # time.sleep(3)
    wait_page_stable(page)

    # 3. 尝试填表 (带重试机制)
    # 只有当还在登录页时才填表
    if LOGIN_URL_KEY in page.url:
        print("[-] 正在寻找输入框...")

        ele_user = None
        # 循环找元素，最多找10秒 (解决 wait.ele 报错问题)
        for _ in range(10):
            ele_user = page.ele('tag:input@@placeholder:请输入账号') or page.ele('#uc-common-account')
            if ele_user: break
            time.sleep(1)

        if ele_user:
            try:
                print("[-] 填写账号密码...")
                ele_pass = page.ele('tag:input@@placeholder:请输入密码') or page.ele('#ucsl-password-edit')
                ele_btn = page.ele('tag:input@@type:submit') or page.ele('#uc-login-submit') or page.ele('text:登录')

                ele_user.clear()
                ele_user.input(username)
                time.sleep(0.5)

                ele_pass.clear()
                ele_pass.input(password)
                time.sleep(0.5)

                if ele_btn:
                    print("[-] 点击登录...")
                    ele_btn.click()
                    time.sleep(3)
            except Exception as e:
                print(f"[!] 填表过程异常: {e}")
        else:
            print("[!] 未找到输入框 (可能已自动登录，或页面未加载)")

    page.get(FULL_TARGET_URL)
    # time.sleep(3)
    wait_page_stable(page)

    # 4. 【死锁等待】
    # 只要 URL 不对，就一直卡在这里提示用户
    print("\n" + "-" * 30)
    print(">>> 等待登录成功 <<<")
    print("脚本正在监控 URL 变化...")
    print(f"目标 URL 特征: {SUCCESS_URL_KEY}")
    print("-" * 30)

    while True:
        current_url = page.url

        # 判定成功条件：包含 关键词
        if (SUCCESS_URL_KEY in current_url) and (LOGIN_URL_KEY not in current_url):
            print(f"\n[√] 捕获到目标 URL: {current_url}")
            print("[-] 登录成功！继续执行任务...")
            break

        # 打印状态防止用户以为死机
        print(f"\r[等待中] 当前: ...{current_url[-30:]} (请手动完成验证)", end="")
        time.sleep(2)

    # 登录成功后，再稳一手，等加载
    wait_page_stable(page)


def switch_to_100_items(page):
    try:
        # ---【刷新和等待】---
        print("[-] 正在刷新页面以初始化状态...")
        page.refresh()
        wait_baidu_report_ready(page)
        print("[-] 检查分页设置...")
        page.scroll.to_bottom()
        time.sleep(1)

        page_btn = page.ele('text:条/页')

        if page_btn:
            current_text = page_btn.text
            if "100" not in current_text:
                print(f"[-] 当前是 [{current_text}]，执行键盘切换...")
                page_btn.click()
                time.sleep(0.5)
                page.actions.type(Keys.DOWN)
                time.sleep(0.2)
                page.actions.type(Keys.DOWN)
                time.sleep(0.2)
                page.actions.type(Keys.ENTER)
                print("[-] 切换指令已发送，等待重新加载...")
                wait_page_stable(page)
            else:
                print("[-] 当前已经是 100条/页")
        else:
            print("[!] 未找到分页按钮")

        page.scroll.to_top()
        time.sleep(0.5)
    except Exception as e:
        print(f"[!] 分页切换出错: {e}")


def switch_device_action(page, device_name):
    """
    Click the promotion device dropdown, then click the option containing device_name
    """
    try:
        print(f"[-] Switching device to: {device_name}...")

        # Click the dropdown containing "promotion device"
        ele = page.ele('text:推广设备')
        if ele:
            ele.click()
            time.sleep(1)

            # Wait for dropdown options and click the one containing device_name
            option = page.ele(f'text:{device_name}')
            if option:
                option.click()
                time.sleep(1)
                print(f"[-] Selected {device_name} device")
            else:
                print(f"[!] Option {device_name} not found")
        else:
            print("[!] Promotion device dropdown not found")

    except Exception as e:
        print(f"[!] Device switch error: {e}")


def set_date_and_capture_via_ui(page, date_param):
    print(f"[-] 准备操作日期: {date_param}")

    # 1. 开启监听 (只听 ReportDataService)
    page.listen.clear()  # 清空旧数据
    page.listen.start("ReportDataService")

    try:
        # 2. 打开日历
        # 尝试多种定位方式，只要有一个能点开就行
        date_trigger = None
        # 找包含年份的文本，或者特定class
        current_year = datetime.datetime.now().year
        date_trigger = page.ele(f'{current_year}/') or page.ele(f'{current_year - 1}/')

        if date_trigger:
            date_trigger.click()
            time.sleep(1)
        else:
            print("[!] 找不到日历触发按钮")
            return None

        # 3. 统一计算目标"日"，并选择最后一个
        target_day_str = ""
        if date_param == "shibao":
            # 如果是今天，计算今天的日期数字
            target_day_str = str(datetime.datetime.now().day)
            print(f"[-] 模式: 今天 (目标数字: {target_day_str})")
        else:
            # 如果是指定日期，提取日
            target_day_str = str(int(date_param.split('-')[-1]))
            print(f"[-] 模式: 指定日期 (目标数字: {target_day_str})")

        # 使用 class 定位所有的日期格子
        all_date_cells = page.eles('css:[class*="one-date-picker-body-month-item"]:not([class*="disabled"])')

        # 筛选出所有文本匹配且可见的格子
        candidates = []
        for cell in all_date_cells:
            print(cell.text.strip(), end=' ')
            if cell.text.strip() == target_day_str:
                if cell.states.is_displayed:
                    candidates.append(cell)
        print()

        if candidates:
            # 选择最后一个 (candidates[-1])
            target_cell = candidates[-1]
            print(f"[-] 找到 {len(candidates)} 个匹配日期，选择最后一个，准备双击...")

            target_cell.click(by_js=True)  # 强制 JS 点击
            print("第一次点击日期")
            time.sleep(0.5)  # 稍微多等一下让页面反应
            target_cell.click(by_js=True)  # 再次强制 JS 点击
            print("第二次点击日期")
        else:
            print(f"[!] 日历上没看到数字 {target_day_str} (已扫描 {len(all_date_cells)} 个格子)")
            return None

        print("[-] UI操作完成，等待并筛选数据包...")
        wait_page_stable(page)

        captured_candidates = []  # 用于暂存抓到的包对象

        # 设置总等待时间（防止网络极其卡顿）
        end_time = time.time() + 30

        print("[-] 进入极速抓包模式 (只存不看)...")

        # --- 第一阶段：极速囤货 ---
        while time.time() < end_time:
            # 使用较短的 timeout (比如2秒)，以便快速响应
            # wait 本身会阻塞，直到有包来或者超时
            res = page.listen.wait(timeout=2)

            if res:
                # 只要 URL 匹配，立刻加入列表，不做任何解析操作！
                if "ReportDataService" in res.url:
                    captured_candidates.append(res)
                    print(f"   > [FAST] 暂存第 {len(captured_candidates)} 个包:  {res.url[-50:]}")

                    # 策略：如果我们已经瞬间抓到了 >=2 个包，大概率目标在里面了
                    # 为了保险，可以再多等一小会，或者如果对速度要求高，抓到2个就撤
                    if len(captured_candidates) >= 2:
                        print("   > [提示] 已捕获足够数量的候选包，停止监听。")
                        break
            else:
                # res 为空说明 timeout 了 (2秒内没新包)
                # 如果我们手里已经有包了，说明那波爆发已经结束了，可以退出了
                if len(captured_candidates) > 0:
                    print("   > [提示] 传输间隙超时，认为传输结束。")
                    break
                # 如果手里没包，说明还没开始传，继续循环等...

        # --- 第二阶段：离线分析 ---
        print(f"[-] 抓包结束，共捕获 {len(captured_candidates)} 个候选包，开始寻找有效数据...")

        for idx, res in enumerate(captured_candidates):
            try:
                # 此时访问 .body 会触发解析，速度慢点也没关系了，因为包已经在内存里了
                body = res.response.body

                # 检查是否为有效数据包
                if isinstance(body, dict) and 'data' in body and 'rows' in body['data']:
                    rows = body['data']['rows']
                    row_count = len(rows)

                    # 打印看一下情况
                    print(f"   [分析包 {idx + 1}] rows数量: {row_count}")

                    # 只要 rows 里有数据，或者 rows 是个列表（即使是空列表也可能是正常响应，看你需求）
                    # 这里假设你需要有数据的包
                    if row_count > 0:
                        print(f"   [√] 锁定目标！在第 {idx + 1} 个包中发现数据。")
                        return body
                    else:
                        # 如果你需要找空数据包（比如当天确实没消费），也可以在这里做逻辑
                        # 暂时假设空包不是我们要找的核心包，除非所有包都空
                        pass

            except Exception as e:
                print(f"   [!] 解析包 {idx + 1} 失败: {e}")

        # 如果循环完了还没找到 >0 的包，但确实有包，可能就是那天没数据
        # 我们可以尝试返回最后一个结构正确的包（即使是空的），或者返回 None
        print("[!] 所有候选包分析完毕，未发现包含 rows>0 的数据。")
        return None

    except Exception as e:
        print(f"[!] UI操作异常: {e}")
        return None


def print_and_save_data(data, device_name="", data_date=""):
    """
    专门用来处理数据打印和保存的函数
    """
    result_list = []  # 初始化空列表

    if data:
        rows = data.get('data', {}).get('rows', [])
        print(f"\n{'-' * 70}")
        print(f"成功获取 {len(rows)} 条数据 (日期: {data_date})")
        print(f"{'账户':<20} | {'设备':<6} | {'展现':<8} | {'点击':<8} | {'消费'}")
        print(f"{'-' * 70}")

        for row in rows:
            if int(row.get('impression', 0)) > 0:
                print(
                    f"{row.get('userName'):<20} | {device_name:<6} | {row.get('impression'):<8} | {row.get('click'):<8} | {row.get('cost')}"
                )

                result_list.append({
                    "date": data_date,
                    "account": row.get('userName'),
                    "device": device_name,
                    "show": int(row.get('impression')),
                    "click": int(row.get('click', 0)),
                    "cost": float(row.get('cost'))
                })

        # --- 将 data_date 传给保存函数 ---
        # save_to_excel(rows, device_name, data_date)
    else:
        print("未获取到数据。")

    return result_list


def baidudata(date_input):
    # 定义结果容器
    all_data_results = []

    # 1. 账号
    username, password = get_credentials_from_excel()
    if not username:
        print("未读取到账号，退出")
        return []

    # 2. 启动
    co = ChromiumOptions()
    co.set_user_data_path(get_chromium_user_data_path())
    co.set_local_port(get_chromium_local_port())
    browser = ChromiumPage(co)

    # --- 获取/切换到具体的标签页 ---
    # 将 browser 传进去，返回具体的 tab 对象给 page 变量
    # 之后所有的 page.ele, page.listen 都是针对这个标签页操作
    page = switch_to_baidu_tab(browser)

    # 3. 严格登录
    strict_login_process(page, username, password)

    # 4. 刷新页面 & 切换 100条/页
    switch_to_100_items(page)

    # 5. 切换设备端口 抓取
    target_date_str = ""
    if date_input == "shibao":
        # === 模式1：如果是 shibao (当天)，只抓一次默认的 ===
        print(f"\n>>> 当前模式: 抓取当天数据 (不切换设备) <<<")
        # target_date_str = datetime.datetime.now().strftime("%Y/%m/%d")
        data = set_date_and_capture_via_ui(page, date_input)
        # 打印并保存
        # print_and_save_data(data, "", target_date_str)
        result_list = print_and_save_data(data, "", target_date_str)
        all_data_results.extend(result_list)

    else:
        # === 模式2：如果是指定日期，分两次抓 ===
        print(f"\n>>> 当前模式: 指定日期 {date_input} (分别抓取 移动 和 计算机) <<<")
        target_date_str = date_input

        # -- 第1轮：移动 --
        switch_device_action(page, "移动")
        data_mobile = set_date_and_capture_via_ui(page, date_input)
        # print_and_save_data(data_mobile, "移动", target_date_str)
        list_mobile = print_and_save_data(data_mobile, "YD", target_date_str)
        all_data_results.extend(list_mobile)

        # -- 第2轮：计算机 --
        switch_device_action(page, "计算机")
        data_pc = set_date_and_capture_via_ui(page, date_input)
        # print_and_save_data(data_pc, "计算机", target_date_str)
        list_pc = print_and_save_data(data_pc, "PC", target_date_str)
        all_data_results.extend(list_pc)

    print(f"\n[完成] 共采集到 {len(all_data_results)} 条有效数据。")
    return all_data_results


# if __name__ == "__main__":
#     baidudata("2026-4-2")
    # baidudata("shibao")
