# CRM 线索数据爬虫
# 功能：抓取 CRM 中的线索数据

import json
import time
import os
import glob
import datetime
import io
import requests
import pandas as pd
from DrissionPage import ChromiumPage, ChromiumOptions
from DrissionPage.common import Keys
from openpyxl import Workbook, load_workbook
from app_config import (
    get_chromium_local_port,
    get_chromium_user_data_path,
    get_env,
    get_platform_credentials,
)

# ================= 配置区域 =================
CONFIG_FILE = "config.xlsx"
RESULT_FILE = "推广数据.xlsx"
TARGET_PLATFORM = "CRM"

# CRM 相关 URL（待用户提供具体地址）
SUCCESS_URL_KEY = get_env("CRM_SUCCESS_URL_KEY")
SIGNOUT_URL_KEY = get_env("CRM_SIGNOUT_URL_KEY")
LOGIN_URL_KEY = get_env("CRM_LOGIN_URL_KEY")
LOGIN_URL = get_env("CRM_LOGIN_URL")
FULL_TARGET_URL = get_env("CRM_TARGET_URL")
CRM_TAB_URL_KEYWORD = get_env("CRM_TAB_URL_KEYWORD")


def switch_to_crm_tab(page):
    """切换到 CRM 标签页（通过URL，包含"duia"）"""
    try:
        tab = page.get_tab(url=CRM_TAB_URL_KEYWORD)
        print(f"[-] 检测到已存在的CRM标签页: {tab.title}，正在激活...")
        tab.set.activate()
        return tab
    except Exception:
        # 未找到，新建标签页
        print("[-] 未检测到CRM标签页，正在新建...")
        new_tab = page.new_tab(FULL_TARGET_URL)
        return new_tab


def get_credentials_from_excel():
    env_username, env_password = get_platform_credentials("CRM")
    if env_username:
        print("[-] Loaded CRM credentials from environment")
        return env_username, env_password

    """从 config.xlsx 读取 CRM 账号密码"""
    if not os.path.exists(CONFIG_FILE):
        return None, None
    try:
        wb = load_workbook(CONFIG_FILE, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row[0] and str(row[0]).strip() == TARGET_PLATFORM:
                print(f"[-] 读取到账号: {row[1]}")
                return str(row[1]), str(row[2])
        return None, None
    except:
        return None, None


def wait_page_stable(page):
    """通用等待函数"""
    try:
        page.wait.doc_loaded(timeout=5)
    except:
        time.sleep(1)
    time.sleep(2)


def strict_login_process(page, username, password):
    """CRM 登录验证流程"""
    wait_page_stable(page)

    print("\n" + "=" * 50)
    print("[-] 开始 CRM 身份验证流程...")

    # 检查是否已登录：URL中有"sso"代表没登录，没有"sso"代表已登录
    current_url = page.url
    is_logged_in = "sso" not in current_url

    if is_logged_in:
        print("[√] 检测到已登录，跳过登录。")
        return

    # 未登录，跳转到登录页面
    print("[-] 未登录，跳转至登录页...")
    page.get(LOGIN_URL)
    wait_page_stable(page)

    # 打印提示，等待手动登录
    print("\n" + "=" * 50)
    print(">>> 需要手动登录 <<<")
    print("请在浏览器中完成登录（输入账号密码、短信验证等）...")
    print("程序将等待，直到检测到登录成功（URL中没有sso）...")
    print("=" * 50)

    while True:
        current_url = page.url

        # 判断成功登录：URL中没有"sso"代表已登录
        if "sso" not in current_url:
            print(f"\n[√] 登录成功！")
            break

        print(f"\r[等待中] {current_url[-60:]}", end="")
        time.sleep(2)

    wait_page_stable(page)


def normalize_date(date_str):
    """
    将日期转换为标准格式 YYYY-MM-DD
    支持: 2026-4-1, 2026-4-01, 2026-04-1, 2026-04-01
    "shibao" 返回当天日期
    """
    if date_str == "shibao":
        return datetime.datetime.now().strftime("%Y-%m-%d")

    if not date_str:
        return date_str

    parts = date_str.split('-')
    if len(parts) != 3:
        return date_str

    year = parts[0]
    month = parts[1].zfill(2)
    day = parts[2].zfill(2)

    return f"{year}-{month}-{day}"


def fetch_crm_data(page, date_input):
    """
    从CRM获取数据
    date_input: "shibao" 或 "2026-04-01"
    返回: DataFrame (直接读取到内存，不落盘)
    """
    # 标准化日期
    target_date = normalize_date(date_input)
    print(f"[-] 目标日期: {target_date}")

    # 1. 跳转到 FULL_TARGET_URL 页面
    print("[-] 跳转到目标页面...")
    page.get(FULL_TARGET_URL)
    wait_page_stable(page)

    # 2. 点击"机会明细"
    print("[-] 点击'机会明细'...")
    chance_detail = page.ele('text:机会明细')
    if chance_detail:
        chance_detail.click()
        time.sleep(3)
    else:
        print("[!] 未找到'机会明细'按钮")
        return

    # 3. 点击日期选择器 (id="wirteTime")
    print("[-] 点击日期选择器...")
    date_picker = page.ele('#wirteTime')
    if date_picker:
        print(f"[-] 找到日期选择器，点击...")
        date_picker.click()
        time.sleep(3)
    else:
        print("[!] 未找到日期选择器")
        return

    # 等待日期选择器弹层出现
    print("[-] 等待日期选择器弹层...")
    time.sleep(3)

    # 尝试通过 JavaScript 查找输入框
    print("[-] 尝试通过 JS 查找日期输入框...")
    js_code = """
    function() {
        var results = {};
        var startInputs = document.querySelectorAll('input[name="daterangepicker_start"]');
        var endInputs = document.querySelectorAll('input[name="daterangepicker_end"]');
        results.count = startInputs.length + ' start, ' + endInputs.length + ' end';
        results.startIds = [];
        results.endIds = [];
        for (var i = 0; i < startInputs.length; i++) {
            results.startIds.push({index: i, visible: startInputs[i].offsetParent !== null, id: startInputs[i].id});
        }
        for (var i = 0; i < endInputs.length; i++) {
            results.endIds.push({index: i, visible: endInputs[i].offsetParent !== null, id: endInputs[i].id});
        }
        return JSON.stringify(results);
    }
    """
    try:
        result = page.run_js(js_code)
        print(f"[-] JS 查找结果: {result}")
    except Exception as e:
        print(f"[!] JS 查找失败: {e}")

    # 4. 使用 JS 输入日期
    print(f"[-] 使用 JS 输入日期: {target_date}")

    js_fill_date = f"""
    function() {{
        var startInputs = document.querySelectorAll('input[name="daterangepicker_start"]');
        var endInputs = document.querySelectorAll('input[name="daterangepicker_end"]');

        // 找到可见的输入框（index 1）
        var visibleStart = null;
        var visibleEnd = null;

        for (var i = 0; i < startInputs.length; i++) {{
            if (startInputs[i].offsetParent !== null) {{
                visibleStart = startInputs[i];
                break;
            }}
        }}

        for (var i = 0; i < endInputs.length; i++) {{
            if (endInputs[i].offsetParent !== null) {{
                visibleEnd = endInputs[i];
                break;
            }}
        }}

        if (visibleStart) {{
            visibleStart.value = '{target_date}';
            visibleStart.dispatchEvent(new Event('input', {{ bubbles: true }}));
            visibleStart.dispatchEvent(new Event('change', {{ bubbles: true }}));
        }}

        if (visibleEnd) {{
            visibleEnd.value = '{target_date}';
            visibleEnd.dispatchEvent(new Event('input', {{ bubbles: true }}));
            visibleEnd.dispatchEvent(new Event('change', {{ bubbles: true }}));
        }}

        return 'start:' + (visibleStart ? 'found' : 'not found') + ', end:' + (visibleEnd ? 'found' : 'not found');
    }}
    """

    try:
        result = page.run_js(js_fill_date)
        print(f"[-] JS 输入结果: {result}")
    except Exception as e:
        print(f"[!] JS 输入失败: {e}")

    # 5. 点击"保存"
    print("[-] 点击'保存'...")

    # 先用 JS 查找并点击
    js_click_save = """
    function() {
        var buttons = document.querySelectorAll('.applyBtn');
        for (var i = 0; i < buttons.length; i++) {
            if (buttons[i].offsetParent !== null && buttons[i].textContent.trim() === '保存') {
                buttons[i].click();
                return 'found and clicked';
            }
        }
        return 'not found';
    }
    """

    try:
        result = page.run_js(js_click_save)
        print(f"[-] JS 点击保存结果: {result}")
    except Exception as e:
        print(f"[!] JS 点击保存失败: {e}")

    time.sleep(1)

    # 备用：用 DrissionPage 点击
    save_btn = page.ele('text:保存')
    if save_btn:
        try:
            save_btn.click(by_js=True)
            print("[-] DrissionPage 点击保存成功")
        except Exception as e:
            print(f"[!] DrissionPage 点击保存失败: {e}")
    else:
        print("[!] 未找到'保存'按钮")

    print("[-] 日期筛选完成，等待数据加载...")
    time.sleep(5)

    # ========== 新流程：市场推广全量数据导出 ==========
    print("\n" + "="*60)
    print(">>> 开始市场推广全量数据导出流程 <<<")
    print("="*60)

    # 6. 选择"机会分类"下拉框，选中"市场推广"
    print("\n[-] 1. 选择'机会分类'下拉框...")
    time.sleep(2)

    # channel2 就是机会分类下拉框，使用 JavaScript 选择"市场推广"
    js_select_channel = """
    function() {
        var select = document.getElementById('channel2');
        if (!select) return 'channel2 not found';

        var options = select.querySelectorAll('option');
        for (var i = 0; i < options.length; i++) {
            if (options[i].innerText.indexOf('市场推广') !== -1) {
                select.value = options[i].value;
                select.dispatchEvent(new Event('change', {bubbles: true}));
                return 'selected: ' + options[i].innerText + ' (value=' + options[i].value + ')';
            }
        }
        return '市场推广 not found in channel2';
    }
    """
    try:
        result = page.run_js(js_select_channel)
        print(f"  选择结果: {result}")
        time.sleep(3)
    except Exception as e:
        print(f"  选择失败: {e}")

    # 7. 点击"搜索"按钮 (id="find")
    print("\n[-] 2. 点击'搜索'按钮...")
    time.sleep(2)
    try:
        js_click_search = """
        function() {
            var btn = document.getElementById('find');
            if (btn) {
                btn.click();
                return 'clicked: ' + btn.innerText;
            }
            return 'find button not found';
        }
        """
        result = page.run_js(js_click_search)
        print(f"  点击搜索结果: {result}")
        time.sleep(5)  # 等待搜索结果
    except Exception as e:
        print(f"[!] 点击搜索失败: {e}")

    # 8. 点击"导出数据"按钮 (id="export")
    print("\n[-] 3. 点击'导出数据'按钮...")
    time.sleep(2)
    try:
        js_click_export = """
        function() {
            var btn = document.getElementById('export');
            if (btn) {
                btn.click();
                return 'clicked export: ' + btn.innerText;
            }
            return 'export button not found';
        }
        """
        result = page.run_js(js_click_export)
        print(f"  点击导出数据结果: {result}")
        time.sleep(3)
    except Exception as e:
        print(f"[!] 点击导出数据失败: {e}")

    # 9. 处理申请成功弹窗，点击"OK"
    print("\n[-] 4. 处理申请成功弹窗...")
    time.sleep(3)
    try:
        # 使用 JavaScript 点击 OK 按钮
        js_click_ok = """
        function() {
            var btns = document.querySelectorAll('button, .btn');
            for (var i = 0; i < btns.length; i++) {
                var text = btns[i].innerText.trim();
                if (text === 'OK' || text === 'ok' || text === '确定') {
                    if (btns[i].offsetParent !== null) {  // 可见
                        btns[i].click();
                        return 'clicked: ' + text;
                    }
                }
            }
            return 'OK button not found';
        }
        """
        result = page.run_js(js_click_ok)
        print(f"  点击OK结果: {result}")
        time.sleep(2)
    except Exception as e:
        print(f"[!] 处理弹窗失败: {e}")

    # 10. 等待进度条完成
    print("\n[-] 5. 等待进度条完成...")
    max_wait = 180  # 最多等待180秒
    start_time = time.time()

    # 智能检测进度条
    js_check_progress = """
    function() {
        // 查找 pace-progress 进度条
        var pace = document.querySelector('.pace-progress');
        if (pace && pace.offsetParent !== null) {
            var style = window.getComputedStyle(pace);
            if (style.display !== 'none') {
                // 获取进度值
                var progressText = pace.getAttribute('data-progress-text') || '';
                return 'pace-progress: ' + progressText;
            }
        }
        // 检查 layui loading
        var layuiLoading = document.querySelector('.layui-layer-loading');
        if (layuiLoading && layuiLoading.offsetParent !== null) {
            return 'layui-loading visible';
        }
        return 'no progress';
    }
    """

    while time.time() - start_time < max_wait:
        time.sleep(3)
        result = page.run_js(js_check_progress)
        if result == 'no progress':
            print(f"  进度条已完成")
            break
        elapsed = int(time.time() - start_time)
        if elapsed % 15 == 0:  # 每15秒打印一次
            print(f"  等待中... ({elapsed}秒): {result}")
        elif elapsed >= 60:
            # 超过60秒后，即使进度条还在显示，也继续执行（可能任务列表已有数据）
            print(f"  等待超时 ({elapsed}秒)，继续执行...")
            break

    time.sleep(3)

    # 11. 点击"查看任务"按钮
    print("\n[-] 6. 点击'查看任务'按钮...")
    time.sleep(2)
    try:
        view_task_btn = page.ele('text:查看任务')
        if view_task_btn:
            view_task_btn.click(by_js=True)
            print("  点击查看任务成功")
            time.sleep(3)  # 等待任务列表弹窗
        else:
            print("[!] 未找到查看任务按钮")
    except Exception as e:
        print(f"[!] 点击查看任务失败: {e}")

    # 12. 在任务列表中获取下载链接
    print("\n[-] 7. 获取下载链接...")
    time.sleep(3)

    # 确保datas目录存在
    datas_dir = './datas'
    if not os.path.exists(datas_dir):
        os.makedirs(datas_dir)
        print(f"  创建目录: {datas_dir}")

    downloaded_file = None
    try:
        # 使用 JavaScript 获取所有下载链接
        js_get_downloads = """
        function() {
            var links = document.querySelectorAll('a');
            var downloads = [];
            for (var i = 0; i < links.length; i++) {
                var href = links[i].href || '';
                var text = links[i].innerText.trim() || '';
                // 找包含"下载"文字或href是xls文件的链接
                if ((text === '下载' && href.indexOf('.xls') !== -1) || href.indexOf('tu.duia.com/upload/excel') !== -1) {
                    downloads.push({
                        'text': text,
                        'href': href
                    });
                }
            }
            return JSON.stringify(downloads);
        }
        """
        result = page.run_js(js_get_downloads)
        import json
        downloads = json.loads(result) if result else []
        print(f"  找到 {len(downloads)} 个下载链接")

        if downloads:
            # 取第一个链接（最新的）
            first_href = downloads[0]['href']
            print(f"  下载链接: {first_href}")

            # 使用 requests 下载文件到内存，不落盘
            try:
                import requests
                response = requests.get(first_href, timeout=60)
                if response.status_code == 200:
                    # 直接从内存读取到 DataFrame
                    file_bytes = io.BytesIO(response.content)
                    df = pd.read_excel(file_bytes, engine='xlrd')
                    print(f"  DataFrame 加载成功，共 {len(df)} 行")

                    # 重命名列：L列是站点ID
                    if len(df.columns) > 11:
                        df.rename(columns={df.columns[11]: 'site_id'}, inplace=True)
                        print(f"  L列(站点ID)已识别，共 {len(df)} 条线索")

                    print("\n[完成] 市场推广数据导出流程执行完毕")
                    return df
                else:
                    print(f"  下载失败，状态码: {response.status_code}")
            except Exception as e:
                print(f"  下载或读取失败: {e}")
        else:
            print("[!] 未找到下载链接")

    except Exception as e:
        print(f"[!] 获取下载链接失败: {e}")

    print("\n[完成] 市场推广数据导出流程执行完毕")
    return pd.DataFrame()


def crmdata(date_input, return_df=True):
    """
    CRM 数据抓取主函数 - 市场推广全量数据导出

    date_input: "shibao" 表示当天，或 "2026-03-31" 这样的指定日期
    return_df: True=返回DataFrame, False=保存到文件(兼容旧逻辑)
    返回: DataFrame(包含L列站点数据) 或 空列表
    """
    # 1. 读取账号
    username, password = get_credentials_from_excel()
    if not username:
        print("未读取到账号，退出")
        return pd.DataFrame()

    # 2. 启动浏览器
    co = ChromiumOptions()
    co.set_user_data_path(get_chromium_user_data_path())
    co.set_local_port(get_chromium_local_port())
    browser = ChromiumPage(co)

    # 切换到 CRM 标签页
    page = switch_to_crm_tab(browser)

    # 3. 登录
    strict_login_process(page, username, password)

    # 4. 执行导出流程，直接返回 DataFrame
    print(f"\n>>> 当前模式: 市场推广全量数据导出 <<<")
    df = fetch_crm_data(page, date_input)

    if df is not None and not df.empty:
        print(f"\n[完成] 市场推广数据已导出，共 {len(df)} 条线索")
    else:
        print(f"\n[!] 未获取到线索数据")

    return df


# if __name__ == "__main__":
    # crmdata("shibao")
    # crmdata("2026-03-31")
