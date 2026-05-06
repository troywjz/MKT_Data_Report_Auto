# 微软必应(Microsoft Bing Ads) 数据爬虫
# 功能：给定日期，返回数据列表
# 已打包成方法 bingdata("shibao")

import json
import re
import time
import os
import datetime
from DrissionPage import ChromiumPage, ChromiumOptions
from DrissionPage.common import Keys
from openpyxl import Workbook, load_workbook
from app_config import (
    get_chromium_local_port,
    get_chromium_user_data_path,
    get_env,
    get_platform_credentials,
)

# ================= 配置区域 =================
CONFIG_FILE = "config.xlsx"
RESULT_FILE = "推广数据.xlsx"
TARGET_PLATFORM = "必应"

# 微软必应广告后台 URL
SUCCESS_URL_KEY = get_env("BING_SUCCESS_URL_KEY")
SIGNOUT_URL_KEY = get_env("BING_SIGNOUT_URL_KEY")
LOGIN_URL_KEY = get_env("BING_LOGIN_URL_KEY")
LOGIN_URL = get_env("BING_LOGIN_URL")
FULL_TARGET_URL = get_env("BING_TARGET_URL")
BING_ACCOUNT_NAME = get_env("BING_ACCOUNT_NAME", "BING_ACCOUNT")


def switch_to_bing_tab(page):
    """切换到必应标签页"""
    target_keyword = "microsoft"
    try:
        tab = page.get_tab(url=target_keyword)
        print(f"[-] 检测到已存在的标签页: {tab.title}，正在激活...")
        tab.set.activate()
        return tab
    except Exception:
        print("[-] 未检测到标签页，正在新建...")
        new_tab = page.new_tab(FULL_TARGET_URL)
        return new_tab


def get_credentials_from_excel():
    env_username, env_password = get_platform_credentials("BING")
    if env_username:
        print("[-] Loaded Bing credentials from environment")
        return env_username, env_password

    """从config.xlsx读取必应账号密码"""
    if not os.path.exists(CONFIG_FILE):
        return None, None
    try:
        wb = load_workbook(CONFIG_FILE, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row[0] and str(row[0]).strip() == TARGET_PLATFORM:
                print(f"[-] 读取到账号: {row[1]}")
                return str(row[1]), str(row[2])
        return None, None
    except:
        return None, None


def wait_bing_report_ready(page, timeout=30):
    """
    等待必应报表页面所有核心元素加载完成（与关系）
    核心元素：日期选择器 + 数据元素（必须同时满足）
    timeout: 最长等待时间(秒)，默认30秒
    """
    # 报表页面的核心元素（页面加载后就应该存在的）
    core_selectors = [
        'css:button.range-display',                    # 日期选择器按钮
        'css:[class*="ba-pc-metric-picker-value"]',  # 数据元素(消费/点击/转化/展示)
    ]

    try:
        page.wait.doc_loaded(timeout=10)
    except:
        pass

    start_time = time.time()
    found = set()

    while time.time() - start_time < timeout:
        for selector in core_selectors:
            if selector not in found:
                try:
                    ele = page.ele(selector, timeout=0.5)
                    if ele:
                        # 数据元素需要检查有实际内容
                        if 'ba-pc-metric-picker-value' in selector:
                            text = ele.text.strip() if ele.text else ''
                            if text:
                                found.add(selector)
                                print(f"  [√] 已加载: {selector}")
                            else:
                                print(f"  [~] 数据元素已出现但内容为空，继续等待...")
                        else:
                            found.add(selector)
                            print(f"  [√] 已加载: {selector}")
                except:
                    pass

        if len(found) == len(core_selectors):
            elapsed = time.time() - start_time
            print(f"[-] 报表页面已就绪 (等待了 {elapsed:.1f} 秒)")
            time.sleep(0.5)
            return True

        time.sleep(0.5)

    print(f"[!] 等待超时({timeout}秒)，已加载: {[s for s in core_selectors if s in found]}")
    time.sleep(2)
    return False


def wait_page_stable(page, timeout=10):
    """
    通用页面等待（仅用于登录页、跳转页等非报表场景）
    只等待文档加载完成，不检查报表特有元素
    """
    try:
        page.wait.doc_loaded(timeout=timeout)
    except:
        pass
    time.sleep(1)


def strict_login_process(page, username, password):
    """登录流程"""
    wait_page_stable(page)

    print("\n" + "=" * 50)
    print("[-] 开始身份验证流程...")

    # 检查是否已登录：如果URL包含SUCCESS_URL_KEY且不包含SIGNOUT_URL_KEY，则已登录
    current_url = page.url
    is_logged_in = SUCCESS_URL_KEY in current_url and (not SIGNOUT_URL_KEY or SIGNOUT_URL_KEY not in current_url)

    if is_logged_in:
        print("[√] 检测到已在报表后台，跳过登录。")
        return

    print("[-] 未登录，跳转至登录页...")
    page.get(LOGIN_URL)
    wait_page_stable(page)

    # 检查是否有已登录用户的选择界面 (newSessionLink)
    print("[-] 检查是否有已登录用户...")
    time.sleep(3)

    # 使用 CSS 选择器检查 newSessionLink 元素
    try:
        new_session_link = page.ele('#newSessionLink')
        if new_session_link:
            print(f"[√] 检测到 newSessionLink 元素，尝试点击...")
            try:
                new_session_link.click()
                print("[√] 点击成功，等待页面跳转...")
                wait_page_stable(page)
                time.sleep(3)

                # 验证登录状态
                current_url = page.url
                if SUCCESS_URL_KEY in current_url and SIGNOUT_URL_KEY not in current_url:
                    print("[√] 已通过用户选择界面登录成功！")
                    return
                else:
                    print("[-] 用户选择后未完全登录，继续等待...")
            except Exception as e:
                print(f"[-] 点击失败: {e}")
        else:
            print("[-] 未检测到 newSessionLink 元素，请手动登录...")
    except Exception as e:
        print(f"[-] 检查 newSessionLink 时出错: {e}")

    # 等待手动登录
    print("\n" + "-" * 30)
    print(">>> 等待手动登录 <<<")
    print(f"目标 URL 特征: {SUCCESS_URL_KEY}")
    print("-" * 30)

    while True:
        current_url = page.url

        # 判断成功登录：包含SUCCESS_URL_KEY且不包含SIGNOUT_URL_KEY
        if SUCCESS_URL_KEY in current_url and (not SIGNOUT_URL_KEY or SIGNOUT_URL_KEY not in current_url):
            print(f"\n[√] 登录成功！")
            break

        # 检测是否在登录页
        if LOGIN_URL_KEY in current_url:
            print(f"\n[!] 检测到登录页面，请手动登录...")

        print(f"\r[等待中] ...{current_url[-30:]}", end="")
        time.sleep(2)

    wait_page_stable(page)


def normalize_date(date_str):
    """
    将日期转换为标准格式 YYYY-MM-DD
    支持: 2026-4-1, 2026-4-01, 2026-04-1, 2026-04-01
    """
    if not date_str or date_str == "shibao":
        return date_str

    parts = date_str.split('-')
    if len(parts) != 3:
        return date_str

    year = parts[0]
    month = parts[1].zfill(2)  # 补齐两位
    day = parts[2].zfill(2)    # 补齐两位

    return f"{year}-{month}-{day}"


def parse_number(text):
    """
    从包含字符的数字字符串中提取数值
    支持格式:
    - "￥1,234.56" -> 1234.56
    - "1.2k" / "1.2K" -> 1200
    - "3.5M" / "3.5m" -> 3500000
    - "1.2万" -> 12000
    - "123" -> 123
    """
    if not text:
        return None

    # 去除常见货币符号、空格和逗号
    cleaned = text.replace('￥', '').replace('¥', '').replace(',', '').replace(' ', '').strip()

    # 检查是否包含字母或万为单位
    multiplier = 1
    has_letter = False

    # 处理中文单位
    if '万' in cleaned:
        cleaned = cleaned.replace('万', '')
        multiplier = 10000
        has_letter = True

    # 处理K/M后缀 (不区分大小写)
    letter_match = re.search(r'([kKmM])$', cleaned)
    if letter_match:
        cleaned = cleaned[:-1]
        unit = letter_match.group(1).lower()
        if unit == 'k':
            multiplier = 1000
        elif unit == 'm':
            multiplier = 1000000
        has_letter = True

    # 尝试提取数字部分
    try:
        num = float(cleaned)
        return num * multiplier
    except ValueError:
        # 如果失败，尝试只提取数字部分
        digit_match = re.search(r'[\d.]+', cleaned)
        if digit_match:
            try:
                return float(digit_match.group()) * multiplier
            except ValueError:
                return None
        return None


def select_date_and_read_data(page, date_param):
    """
    选择日期后，直接从页面读取花费、点击、展示数据
    通过键盘输入方式选择日期
    """
    # shibao 模式不需要解析日期
    date_input = None
    if date_param != "shibao":
        # 标准化日期格式 (2026-04-01 -> 2026/4/1)
        date_param = normalize_date(date_param)
        # 转换为 2026/4/1 格式
        date_formatted = '/'.join(date_param.split('-'))
        # 去掉前导零
        parts = date_formatted.split('/')
        date_input = f"{parts[0]}/{int(parts[1])}/{int(parts[2])}"
        print(f"[-] 准备操作日期: {date_param}")
        print(f"[-] 输入格式: {date_input}")
    else:
        print(f"[-] 模式: 今天 (shibao)")

    # 关闭可能弹出的弹窗
    try:
        page.ele('css:body').click()
        time.sleep(0.3)
    except:
        pass

    # 打开日期选择器
    date_trigger = page.ele('css:button.range-display')
    if not date_trigger:
        date_trigger = page.ele('text:昨天:')

    if not date_trigger:
        print("[!] 找不到日期选择器")
        return None

    print("[-] 找到日期选择器，点击...")
    date_trigger.click()
    time.sleep(1)

    if date_param == "shibao":
        # 今天 - 直接点击今天按钮
        print("[-] 模式: 今天")
        today_btn = page.ele('text:今天') or page.ele('text:Today')
        if today_btn:
            today_btn.click()
            time.sleep(2)
    else:
        # 指定日期 - 使用键盘输入
        print(f"[-] 模式: 指定日期 (输入: {date_input})")

        # 等待日期选择器弹窗完全打开
        time.sleep(1.5)

        # 尝试查找日期输入框
        start_input = page.ele('css:input[aria-label*="开始"]') or page.ele('css:input[placeholder*="开始"]') or page.ele('css:input[aria-label*="Start"]')
        end_input = page.ele('css:input[aria-label*="结束"]') or page.ele('css:input[placeholder*="结束"]') or page.ele('css:input[aria-label*="End"]')

        print(f"[-] 开始日期输入框: {start_input}")
        print(f"[-] 结束日期输入框: {end_input}")

        if start_input and end_input:
            print("[-] 找到日期输入框，直接输入...")

            # 输入开始日期
            print(f"[-] 输入开始日期: {date_input}")
            start_input.click()
            time.sleep(0.5)
            # 使用 Ctrl+A 全选，然后直接输入新值
            page.actions.key_down(Keys.CONTROL)
            page.actions.type('a')
            page.actions.key_up(Keys.CONTROL)
            time.sleep(0.5)
            start_input.input(date_input)
            time.sleep(1)

            # 直接点击结束日期输入框切换焦点
            print("[-] 点击结束日期输入框...")
            end_input = page.ele('css:input[aria-label*="结束"]')
            if end_input:
                end_input.click()
                time.sleep(1)
            else:
                print("[!] 未找到结束日期输入框")
            time.sleep(1)

            # 输入结束日期
            print(f"[-] 输入结束日期: {date_input}")
            end_input = page.ele('css:input[aria-label*="结束"]')
            if end_input:
                end_input.click()
                time.sleep(0.5)
                page.actions.key_down(Keys.CONTROL)
                page.actions.type('a')
                page.actions.key_up(Keys.CONTROL)
                time.sleep(0.5)
                end_input.input(date_input)
                time.sleep(0.5)
                # 按回车确认输入
                page.actions.input(Keys.ENTER)
            else:
                print("[!] 结束日期输入框获取失败")
            time.sleep(1)
        else:
            print("[-] 未找到独立输入框，使用键盘操作...")
            # 全选开始日期并输入
            print("[-] 全选开始日期...")
            page.actions.key_down(Keys.CONTROL)
            time.sleep(0.3)
            page.actions.type('a')
            time.sleep(0.3)
            page.actions.key_up(Keys.CONTROL)
            time.sleep(0.3)
            page.actions.input(Keys.DELETE)
            time.sleep(0.5)

            # 输入开始日期
            print(f"[-] 输入开始日期: {date_input}")
            page.actions.type(date_input)
            time.sleep(1)

            # 按 Tab 切换到结束日期
            print("[-] 按 Tab 切换到结束日期...")
            page.actions.input(Keys.TAB)
            time.sleep(1)

            # 全选结束日期并删除
            print("[-] 全选结束日期...")
            page.actions.key_down(Keys.CONTROL)
            time.sleep(0.3)
            page.actions.type('a')
            time.sleep(0.3)
            page.actions.key_up(Keys.CONTROL)
            time.sleep(0.3)
            page.actions.input(Keys.DELETE)
            time.sleep(0.5)

            # 输入结束日期
            print(f"[-] 输入结束日期: {date_input}")
            page.actions.type(date_input)
            time.sleep(0.5)
            # 按回车确认输入
            page.actions.input(Keys.ENTER)
            time.sleep(1)

        # 点击应用按钮确认
        print("[-] 查找并点击应用按钮...")
        time.sleep(0.5)
        apply_btn = page.ele('text:应用') or page.ele('text:Apply')
        print(f"[-] 应用按钮: {apply_btn}")

        if apply_btn:
            try:
                apply_btn.click(by_js=True)
                print("[-] 应用按钮点击成功")
            except Exception as e:
                print(f"[!] 应用按钮点击失败: {e}")
                page.actions.input(Keys.ENTER)
        else:
            print("[-] 未找到应用按钮，按 Enter 确认...")
            page.actions.input(Keys.ENTER)

        time.sleep(3)

    # 关闭可能残留的日历
    try:
        page.ele('css:body').click()
        time.sleep(0.5)
    except:
        pass

    # 从页面读取数据
    print("[-] 从页面读取数据...")
    data = read_bing_data_from_page(page)

    return data


def read_bing_data_from_page(page):
    """
    从必应页面读取花费、点击、转化、展示数据
    通过 class="fui-Text ba-pc-metric-picker-value" 定位元素
    """
    result_list = []

    try:
        # 等待数据加载
        time.sleep(2)

        # 滚动到顶部
        page.scroll.to_top()
        time.sleep(1)

        print("[-] 开始搜索页面数据...")

        # 找到所有包含 ba-pc-metric-picker-value 的元素
        metric_elements = page.eles('css:[class*="ba-pc-metric-picker-value"]')

        print(f"[-] 找到 {len(metric_elements)} 个数据元素")

        data_values = {}

        # 根据用户反馈：元素顺序是 [消费, 点击, 转化, 展现]
        # 使用索引顺序识别
        type_mapping = {
            0: 'cost',      # 第一个是消费
            1: 'clicks',    # 第二个是点击
            2: 'conversions', # 第三个是转化
            3: 'impressions' # 第四个是展现
        }

        for i, ele in enumerate(metric_elements):
            try:
                text = ele.text.strip()
                aria_label = ele.attr('aria-label') or ''
                print(f"  [{i}] text: {text}, aria-label: {aria_label}")

                # 按索引顺序识别
                if i in type_mapping:
                    data_type = type_mapping[i]

                    # 提取数字
                    num = parse_number(text)

                    if num is not None:
                        if data_type == 'cost':
                            data_values['cost'] = float(num)
                            print(f"    -> 花费: {num}")
                        else:
                            # 点击、转化、展现是整数
                            if data_type == 'clicks':
                                data_values['clicks'] = int(num)
                                print(f"    -> 点击: {num}")
                            elif data_type == 'conversions':
                                data_values['conversions'] = int(num)
                                print(f"    -> 转化: {num}")
                            elif data_type == 'impressions':
                                data_values['impressions'] = int(num)
                                print(f"    -> 展示: {num}")

            except Exception as e:
                print(f"  [!] 处理元素 {i} 失败: {e}")

        print(f"\n[-] 获取到的数据: {data_values}")

        # 如果没有通过class找到，尝试通过aria-label搜索
        if not data_values:
            print("[-] 尝试通过 aria-label 搜索...")
            all_spans = page.eles('tag:span')
            for span in all_spans:
                try:
                    text = span.text.strip()
                    aria = span.attr('aria-label') or ''

                    # 检查是否包含金额（花费）
                    if '￥' in text or '￥' in aria:
                        num = parse_number(text)
                        if num is not None:
                            data_values['cost'] = float(num)
                            print(f"  找到花费: {num}")
                    # 检查是否包含纯数字
                    else:
                        num = parse_number(text)
                        if num is not None and len(text) > 2:
                            # 尝试判断是点击还是展示
                            if 'impressions' in str(span).lower() or '展示' in str(span):
                                data_values['impressions'] = int(num)
                            elif 'clicks' in str(span).lower() or '点击' in str(span):
                                data_values['clicks'] = int(num)
                except:
                    pass

        # 返回提取的数据
        if data_values:
            return {
                "data": [{
                    "account": BING_ACCOUNT_NAME,
                    "show": data_values.get('impressions', 0),
                    "click": data_values.get('clicks', 0),
                    "cost": data_values.get('cost', 0.0),
                    "conversions": data_values.get('conversions', 0)
                }],
                "source": "page_read"
            }
        else:
            print("[!] 未找到任何数据")
            return None

    except Exception as e:
        print(f"[!] 读取数据失败: {e}")
        import traceback
        traceback.print_exc()
        return None


def print_and_save_data(data, device_name="", data_date=""):
    """处理必应数据打印和保存"""
    result_list = []

    if data:
        rows = data.get('data', [])
        if not rows:
            print("[!] 未获取到数据行")
            return result_list

        print(f"\n{'-' * 70}")
        print(f"成功获取 {len(rows)} 条数据 (日期: {data_date})")
        print(f"{'账户':<20} | {'设备':<6} | {'展现':<8} | {'点击':<8} | {'消费'}")
        print(f"{'-' * 70}")

        for row in rows:
            account_name = BING_ACCOUNT_NAME
            impressions = int(row.get('show', 0))
            clicks = int(row.get('click', 0))
            cost = float(row.get('cost', 0))

            if cost > 0 or clicks > 0:
                print(f"{account_name:<20} | {device_name:<6} | {impressions:<8} | {clicks:<8} | {cost:.2f}")
                result_list.append({
                    "date": data_date,
                    "account": account_name,
                    "device": device_name,
                    "show": impressions,
                    "click": clicks,
                    "cost": cost
                })
    else:
        print("未获取到数据。")

    return result_list


def bingdata(date_input):
    """
    必应数据抓取主函数
    date_input: "shibao" 表示当天，或 "2026-03-31" 这样的指定日期
    """
    all_data_results = []

    # 1. 账号
    username, password = get_credentials_from_excel()
    if not username:
        print("未读取到账号，退出")
        return []

    # 2. 启动浏览器
    co = ChromiumOptions()
    co.set_user_data_path(get_chromium_user_data_path())
    co.set_local_port(get_chromium_local_port())
    browser = ChromiumPage(co)

    # 切换到必应标签页
    page = switch_to_bing_tab(browser)

    # 3. 登录
    strict_login_process(page, username, password)

    # 刷新页面
    page.refresh()
    wait_bing_report_ready(page)

    # 4. 选择日期并读取数据
    print(f"\n>>> 当前模式: 抓取数据 <<<")

    device_name = "" if date_input == "shibao" else "PC"

    try:
        data = select_date_and_read_data(page, date_input)
    except Exception as e:
        print(f"[!] 选择日期失败: {e}")
        import traceback
        traceback.print_exc()
        data = None

    try:
        result_list = print_and_save_data(data, device_name, date_input)
        all_data_results.extend(result_list)
    except Exception as e:
        print(f"[!] 处理数据失败: {e}")
        import traceback
        traceback.print_exc()

    print(f"\n[完成] 共采集到 {len(all_data_results)} 条有效数据。")
    return all_data_results


# if __name__ == "__main__":
#     bingdata("shibao")
#     bingdata("2026-4-16")
