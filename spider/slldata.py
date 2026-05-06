# 360点睛 - 超级账号 数据爬虫
# 功能：给定日期，返回数据列表
# 已打包成方法 360data("shibao")

# 调试时，打开最后几行的 if __name__ == "__main__": ...

import json
import time
import os
import datetime
from DrissionPage import ChromiumPage, ChromiumOptions
from DrissionPage.common import Keys
from openpyxl import Workbook, load_workbook
from app_config import (
    get_chromium_local_port,
    get_chromium_user_data_path,
    get_env,
    get_platform_credentials,
)

# ================= 配置区域 =================
CONFIG_FILE = "config.xlsx"
RESULT_FILE = "推广数据.xlsx"
TARGET_PLATFORM = "三六零"

# 只有当 URL 包含这个时，才算真正进入后台
SUCCESS_URL_KEY = get_env("SLL_SUCCESS_URL_KEY")
LOGIN_URL_KEY = get_env("SLL_LOGIN_URL_KEY")
LOGIN_URL = get_env("SLL_LOGIN_URL")
FULL_TARGET_URL = get_env("SLL_TARGET_URL")



# date_input = "2025-11-24"
# date_input = "shibao"

# ===========================================

def switch_to_platform_tab(page):
    """
    寻找并切换到包含 平台名 的标签页
    如果没有，则新建一个
    """
    # 尝试查找 URL 包含 keyword 的标签页
    target_keyword = "360"

    try:
        # 尝试查找标签页
        # 如果找不到，DrissionPage 会直接报错，跳转到 except
        tab = page.get_tab(url=target_keyword)

        print(f"[-] 检测到已存在的标签页: {tab.title}，正在激活...")
        tab.set.activate()
        return tab

    except Exception:
        # 捕获找不到标签页的错误，执行新建操作
        print("[-] 未检测到标签页，正在新建...")
        new_tab = page.new_tab(FULL_TARGET_URL)
        return new_tab

def get_credentials_from_excel():
    env_username, env_password = get_platform_credentials("SLL")
    if env_username:
        print("[-] Loaded SLL credentials from environment")
        return env_username, env_password

    if not os.path.exists(CONFIG_FILE): return None, None
    try:
        wb = load_workbook(CONFIG_FILE, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row[0] and str(row[0]).strip() == TARGET_PLATFORM:
                print(f"[-] 读取到账号: {row[1]}")
                return str(row[1]), str(row[2])
        return None, None
    except:
        return None, None


def save_to_excel(data_rows, device_name="", data_date=""):
    try:
        if os.path.exists(RESULT_FILE):
            wb = load_workbook(RESULT_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["数据日期", "账户名称", "设备", "展现", "点击", "消费"])

        count = 0
        for row in data_rows:
            if int(row.get('impression', 0)) > 0:
                line = [
                    data_date,
                    row.get('userName'),
                    device_name,
                    int(row.get('impression')),
                    int(row.get('click', 0)),
                    float(row.get('cost'))
                ]
                ws.append(line)
                count += 1
        wb.save(RESULT_FILE)
        print(f"[√] 已将 {count} 条数据写入 {RESULT_FILE}")
    except Exception as e:
        print(f"[!] 写入 Excel 失败: {e}")


def wait_360_report_ready(page, timeout=30):
    """
    等待360报表页面所有核心元素加载完成（与关系）
    核心元素：分页控件(条/页) + 日期选择器
    timeout: 最长等待时间(秒)，默认30秒
    """
    core_selectors = [
        'text:条/页',                                  # 分页控件
        'css:.el-date-editor.el-range-editor',       # 日期选择器
    ]

    try:
        page.wait.doc_loaded(timeout=10)
    except:
        pass

    start_time = time.time()
    found = set()

    while time.time() - start_time < timeout:
        for selector in core_selectors:
            if selector not in found:
                try:
                    ele = page.ele(selector, timeout=0.5)
                    if ele:
                        found.add(selector)
                        print(f"  [√] 已加载: {selector}")
                except:
                    pass

        if len(found) == len(core_selectors):
            elapsed = time.time() - start_time
            print(f"[-] 报表页面已就绪 (等待了 {elapsed:.1f} 秒)")
            time.sleep(0.5)
            return True

        time.sleep(0.5)

    print(f"[!] 等待超时({timeout}秒)，已加载: {[s for s in core_selectors if s in found]}")
    time.sleep(2)
    return False


def wait_page_stable(page):
    """通用等待函数（仅用于登录页等非报表场景）"""
    try:
        page.wait.doc_loaded(timeout=5)
    except:
        time.sleep(1)
    time.sleep(2)


def strict_login_process(page, username, password):
    """
    【严厉版】登录流程
    如果不进后台，死都不出来
    """
    wait_page_stable(page)

    print("\n" + "=" * 50)
    print("[-] 开始身份验证流程...")

    # 1. 判断当前状态
    # 如果不包含 关键词，说明没登录
    is_logged_in = SUCCESS_URL_KEY in page.url

    if is_logged_in:
        print("[√] 检测到已在报表后台，跳过登录。")
        return

    # 2. 进入登录页
    print("[-] 未登录，跳转至登录页...")
    # page.get(FULL_TARGET_URL)
    # time.sleep(3)
    page.get(LOGIN_URL)
    wait_page_stable(page)

    # 3. 尝试填表 (带重试机制)
    # 只有当还在登录页时才填表
    # if LOGIN_URL_KEY in page.url:
    if page.url == LOGIN_URL:
        print("[-] 正在寻找输入框...")

        ele_user = None
        # 循环找元素，最多找10秒 (解决 wait.ele 报错问题)
        for _ in range(10):
            ele_user = page.ele('tag:input@@placeholder:请输入账号') or page.ele('#uc-common-account')
            if ele_user: break
            time.sleep(1)

        if ele_user:
            try:
                print("[-] 填写账号密码...")
                ele_pass = page.ele('tag:input@@placeholder:请输入密码') or page.ele('#ucsl-password-edit')
                ele_btn = page.ele('tag:input@@type:submit') or page.ele('#uc-login-submit') or page.ele('text:登录')

                ele_user.clear()
                ele_user.input(username)
                time.sleep(0.5)

                ele_pass.clear()
                ele_pass.input(password)
                time.sleep(0.5)

                if ele_btn:
                    print("[-] 点击登录...")
                    ele_btn.click()
                    time.sleep(3)
            except Exception as e:
                print(f"[!] 填表过程异常: {e}")
        else:
            print("[!] 未找到输入框 (可能已自动登录，或页面未加载)")

    page.get(FULL_TARGET_URL)
    # time.sleep(3)
    wait_page_stable(page)

    # 4. 【死锁等待】
    # 只要 URL 不对，就一直卡在这里提示用户
    print("\n" + "-" * 30)
    print(">>> 等待登录成功 <<<")
    print("脚本正在监控 URL 变化...")
    print(f"目标 URL 特征: {SUCCESS_URL_KEY}")
    print("-" * 30)

    while True:
        current_url = page.url

        # 判定成功条件：包含 关键词
        if SUCCESS_URL_KEY in current_url:
            print(f"\n[√] 捕获到目标 URL: {current_url}")
            print("[-] 登录成功！继续执行任务...")
            break

        # 打印状态防止用户以为死机
        print(f"\r[等待中] 当前: ...{current_url[-30:]} (请手动完成验证)", end="")
        time.sleep(2)

    # 登录成功后，再稳一手，等加载
    wait_page_stable(page)


def switch_to_100_items(page):
    """
    360分页设置：每页30条 → 切换到100条
    使用CSS class定位更准确
    """
    try:
        print("[-] 正在刷新页面以初始化状态...")
        page.refresh()
        wait_360_report_ready(page)
        print("[-] 检查分页设置...")

        # 先滚动到页面底部找分页控件
        page.scroll.to_bottom()
        time.sleep(1)

        # 方式1：通过class定位el-input__inner的input，placeholder包含"请选择"
        page_btn = page.ele('css:el-input__suffix')

        # 方式2：如果上面找不到，尝试通过父元素的class定位
        if not page_btn:
            page_btn = page.ele('css:.el-select')

        # 方式3：尝试找包含"条/页"的元素
        if not page_btn:
            page_btn = page.ele('text:条/页')

        if page_btn:
            print(f"[-] 找到分页控件，点击...")
            page_btn.click()
            time.sleep(0.5)

            # 找到下拉选项中的100条
            option_100 = page.ele('text:100条/页')
            if option_100:
                option_100.click()
                print("[-] 已切换到每页100条")
                wait_page_stable(page)
            else:
                print("[-] 未找到100条选项，当前可能是默认设置")
        else:
            print("[!] 未找到分页按钮")

        page.scroll.to_top()
        time.sleep(0.5)
    except Exception as e:
        print(f"[!] 分页切换出错: {e}")


def switch_device_action(page, down_times, device_name):
    """
    点击推广设备下拉框，按 down_times 次下键，然后回车
    """
    try:
        print(f"[-] 正在切换设备到: {device_name} (按键 {down_times} 次)...")

        # 定位包含"推广设备"的元素，点击它
        # 注意：百度后台这里可能是一个下拉框，文本包含"推广设备"
        ele = page.ele('text:推广设备')
        if ele:
            ele.click()
            time.sleep(0.5)

            # 按 N 次下键
            for _ in range(down_times):
                page.actions.type(Keys.DOWN)
                time.sleep(0.2)

            # 按回车确认
            page.actions.type(Keys.ENTER)

            # 等待加载
            print("[-] 设备切换指令发送，等待数据加载...")
            wait_page_stable(page)
        else:
            print("[!] 未找到'推广设备'下拉框，无法切换设备")

    except Exception as e:
        print(f"[!] 切换设备异常: {e}")


def set_date_and_capture_via_ui(page, date_param):
    print(f"[-] 准备操作日期: {date_param}")

    # 360的API关键字
    page.listen.clear()
    page.listen.start("api")

    try:
        # ========================================
        # 第一步：定位日期选择器 - 使用class精确定位
        # ========================================
        date_trigger = None

        # 方式1：使用class精确定位第二个日期选择器（daterange范围选择）
        date_trigger = page.ele('css:.el-date-editor.el-range-editor.mr10')

        # 方式2：如果上面找不到，尝试通过class包含daterange
        if not date_trigger:
            date_trigger = page.ele('css:[class*="daterange"]')

        # 方式3：尝试找所有日期编辑器，选择第二个
        if not date_trigger:
            all_date_editors = page.eles('css:.el-date-editor')
            if len(all_date_editors) >= 2:
                date_trigger = all_date_editors[1]  # 取第二个
                print(f"[-] 使用第2个日期编辑器")

        if not date_trigger:
            print("[!] 找不到日期选择器")
            return None

        print(f"[-] 找到日期选择器，点击...")
        date_trigger.click()
        time.sleep(1)

        # ========================================
        # 第二步：选择日期
        # ========================================
        if date_param == "shibao":
            # 模式：今天（时报）
            today_btn = page.ele('text:今天')
            if today_btn:
                print("[-] 点击'今天'按钮")
                today_btn.click()
                time.sleep(1)
            else:
                # 尝试按Enter确认默认选中的今天
                page.actions.type(Keys.ENTER)
                time.sleep(0.5)
        else:
            # 模式：指定日期（日报）
            target_date = date_param
            target_day = int(date_param.split('-')[-1])
            target_month = date_param.split('-')[1]
            print(f"[-] 模式: 指定日期 (目标: {target_date}, 日: {target_day})")

            # 找到日期格子
            all_date_cells = page.eles('css:.el-date-table td:not(.disabled):not(.prev):not(.next)')
            print(f"[-] 扫描到 {len(all_date_cells)} 个日期格子")

            # 打印可见的日期
            visible_dates = []
            for cell in all_date_cells[:15]:
                try:
                    txt = cell.text.strip()
                    if txt and txt.isdigit():
                        visible_dates.append(txt)
                        print(f"  {txt}", end=" ")
                except:
                    pass
            print()

            # 找到目标日期并点击
            candidates = [c for c in all_date_cells if c.text.strip() == str(target_day)]
            if candidates:
                print(f"[-] 点击日期: {target_day}")
                candidates[-1].click(by_js=True)
                time.sleep(0.3)

                # 对于日期范围选择器，点击第二次选择结束日期
                candidates[-1].click(by_js=True)
                time.sleep(0.3)

                # 等待日历关闭后，再点击查询
                time.sleep(0.5)
            else:
                print(f"[!] 未找到日期 {target_day}")

        # ========================================
        # 第三步：点击查询按钮
        # ========================================
        query_btn = page.ele('text:查询')
        if query_btn:
            print("[-] 点击查询按钮")
            query_btn.click()
            time.sleep(2)
        else:
            print("[!] 未找到查询按钮")

        print("[-] 等待数据包...")

        captured_candidates = []  # 用于暂存抓到的包对象

        # 设置总等待时间（防止网络极其卡顿）
        end_time = time.time() + 30

        print("[-] 进入极速抓包模式 (只存不看)...")

        # --- 第一阶段：极速囤货 ---
        # 360的API可能包含多种关键字，使用更宽松的匹配
        while time.time() < end_time:
            # 使用较短的 timeout (比如2秒)，以便快速响应
            res = page.listen.wait(timeout=2)

            if res:
                # 360的数据包关键字可能是以下任一关键词
                # - api (最常见)
                # - report
                # - data
                # - mcc (360营销中心的缩写)
                # 使用更宽松的匹配：url中包含这些关键词之一即可
                url_lower = res.url.lower()
                if any(kw in url_lower for kw in ['api', 'report', 'data', 'mcc']):
                    captured_candidates.append(res)
                    print(f"   > [FAST] 暂存第 {len(captured_candidates)} 个包:  {res.url[-50:]}")

                    # 360可能一次性返回所有数据抓到1个就够
                    if len(captured_candidates) >= 1:
                        print("   > [提示] 已捕获数据包，停止监听。")
                        break
            else:
                if len(captured_candidates) > 0:
                    print("   > [提示] 传输间隙超时，认为传输结束。")
                    break

        # --- 第二阶段：分析数据包 ---
        print(f"[-] 抓包结束，共捕获 {len(captured_candidates)} 个候选包，开始分析...")

        for idx, res in enumerate(captured_candidates):
            try:
                body = res.response.body
                print(f"\n--- 分析包 {idx + 1} ---")

                if isinstance(body, dict):
                    print(f"Keys: {list(body.keys())}")
                    data = body.get('data', [])

                    if isinstance(data, list) and len(data) > 0:
                        first_item = data[0]
                        print(f"数据条数: {len(data)}")
                        print(f"第一条数据keys: {list(first_item.keys())}")

                        # 检查是否有展现和点击字段
                        if 'impression' in first_item or 'click' in first_item:
                            print(">>> 找到包含展现/点击的数据！")
                            return body
                        else:
                            print(">>> 该包没有展现点击字段")

            except Exception as e:
                print(f"   [!] 解析包 {idx + 1} 失败: {e}")

        # 返回第一个有数据的包
        for res in captured_candidates:
            try:
                body = res.response.body
                if isinstance(body, dict) and body.get('data'):
                    return body
            except:
                pass

        return None

    except Exception as e:
        print(f"[!] UI操作异常: {e}")
        return None


def print_and_save_data(data, device_name="", data_date=""):
    """
    处理360数据打印和保存
    360返回的数据格式: {"data": [...], "count": N}
    数据项包含: uname, views(展现), clicks(点击), costs(消费)等
    """
    result_list = []

    if data:
        rows = data.get('data', [])
        if not rows:
            rows = data.get('data', {}).get('rows', [])

        print(f"\n{'-' * 70}")
        print(f"成功获取 {len(rows)} 条数据 (日期: {data_date})")
        print(f"{'账户':<20} | {'设备':<6} | {'展现':<8} | {'点击':<8} | {'消费'}")
        print(f"{'-' * 70}")

        for row in rows:
            # 360字段映射
            # uname/pname -> 账户名
            # views -> 展现
            # clicks -> 点击
            # costs -> 消费（直接使用）

            account_name = row.get('uname', '') or row.get('pname', '')

            # 获取展现和点击
            views = int(row.get('views', 0) or 0)
            clicks = int(row.get('clicks', 0) or 0)

            # 只保留 costs 字段（或者 search_cost）
            cost = float(row.get('costs', 0) or 0)
            if cost == 0:
                cost = float(row.get('search_cost', 0) or 0)

            # 只要有消费或者有点击就记录
            if cost > 0 or clicks > 0:
                print(f"{account_name:<20} | {device_name:<6} | {views:<8} | {clicks:<8} | {cost:.2f}")

                result_list.append({
                    "date": data_date,
                    "account": account_name,
                    "device": device_name,
                    "show": views,
                    "click": clicks,
                    "cost": cost
                })

    else:
        print("未获取到数据。")

    return result_list


def slldata(date_input):
    # 定义结果容器
    all_data_results = []

    # 1. 账号
    username, password = get_credentials_from_excel()
    if not username:
        print("未读取到账号，退出")
        return []

    # 2. 启动
    co = ChromiumOptions()
    co.set_user_data_path(get_chromium_user_data_path())
    co.set_local_port(get_chromium_local_port())
    browser = ChromiumPage(co)

    # --- 获取/切换到具体的标签页 ---
    # 将 browser 传进去，返回具体的 tab 对象给 page 变量
    # 之后所有的 page.ele, page.listen 都是针对这个标签页操作
    page = switch_to_platform_tab(browser)

    # 3. 严格登录
    strict_login_process(page, username, password)

    # 4. 刷新页面 & 切换 100条/页
    switch_to_100_items(page)

    # 5. 抓取数据（360不需要分设备，直接查询即可）
    target_date_str = ""
    device_name = "" if date_input == "shibao" else "PC"
    print(f"\n>>> 当前模式: 抓取数据 <<<")

    data = set_date_and_capture_via_ui(page, date_input)
    result_list = print_and_save_data(data, device_name, target_date_str)
    all_data_results.extend(result_list)

    print(f"\n[完成] 共采集到 {len(all_data_results)} 条有效数据。")
    return all_data_results


# if __name__ == "__main__":
#     slldata("2026-3-30")
    # slldata("shibao")
